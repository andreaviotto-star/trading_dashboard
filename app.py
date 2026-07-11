"""
Quant Trading System Analytics Dashboard  v11.0
=================================================
Usage: streamlit run app.py

═══════════════════════════════════════════════════════════════
CHANGES IN v11.0
═══════════════════════════════════════════════════════════════

1. DYNAMIC SIZING (Tab 1)
   - Per-system number_input in Tab 1 System Explorer.
   - Sizing panel above the overview table: compact number_inputs
     for all systems in a responsive grid.
   - Default = qty from ReadMe.txt; user can override.
   - All metrics and equity curves rescale live:
       scaled_P&L = raw_P&L × (current_n / default_n)
   - "Reset to defaults" button clears session_state overrides.
   - Raw P&L from TradeStation is NEVER changed — only displayed
     P&L is multiplied by the scaling ratio.

2. SYSTEM NAMES FROM README.TXT (everywhere)
   - si["display_name"] = name exactly as written left of ":" in ReadMe.
   - 16 / 18 files matched; 2 files absent from ReadMe show
     their filename + ⚠️ badge.
   - ReadMe names used in: overview table, equity legend,
     correlation heatmap labels, portfolio bar chart, Excel export.

3. SIZING DEFAULTS FROM README.TXT (improved parsing)
   - Regex r'(N+)\\s*([A-Z]{2,3})' on text before first comma.
   - ES reversal: first qty = 3 (conservative L2), label "3+4 MES (L2+L3)".
   - GC donchian: qty = 1 GC (strips "prima erano 5 MGC…" note).
   - NQ breakout sessione continua: 1 NQ (full contract, not micro).
   - CL donchian: 1 CL (full contract).

═══════════════════════════════════════════════════════════════
RETAINED FROM v10.0
═══════════════════════════════════════════════════════════════
  • Raw-trade P&L engine (no TS multiplier errors)
  • Light theme (plotly_white)
  • 10-year default lookback
  • Risk Parity sizing (Tab 3)
  • Drawdown Decomposition (Tab 2)
  • Clustered Correlation Heatmap (Tab 3)
  • Excel Export (sidebar)
  • What-If Sizing Simulator (Tab 2)
  • Monte Carlo Projection (Tab 4)
  • Regime Analysis (Tab 4)
  • Rolling Health Monitor (Tab 4)

COLUMN MAP — "Trades List" tab (verified all 18 xlsx):
  Entry row: col[0]=trade#, col[1]=direction, col[2]=entry_dt,
             col[6]=n_contracts, col[7]=trade_pnl
  Exit row:  col[0]=None, col[1]=exit_type, col[2]=exit_dt,
             col[6]=trade P&L ($, full position) ← canonical
             col[7]=cumulative P&L ($)
  Dollar values = full-position dollars from TradeStation.
  NEVER multiply by contract count again.
"""

import io
import os
import re
import warnings
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from scipy.cluster.hierarchy import linkage, leaves_list
from scipy.spatial.distance import squareform

warnings.filterwarnings("ignore")

# ── Trade-matching engine (new, v11.1) — imported defensively so the rest of
# the dashboard keeps working even if the module or its extra dependency
# (pdfplumber, only needed for the optional PDF reconciliation) is missing.
try:
    import tradestation_matcher as tsm
    _TSM_IMPORT_ERROR = None
except Exception as _tsm_exc:
    tsm = None
    _TSM_IMPORT_ERROR = str(_tsm_exc)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
DATA_DIR = Path(os.path.dirname(os.path.abspath(__file__)))

# $/contract round-trip commission (NOT flat per-trade) -- see build_net_equity().
# Defaults below are blended from Andrea's real June 2026 statement, weighted
# by actual contract volume (not a simple average of per-root rates):
#   micro: (MES $401.94 + MGC $29.28 + MNQ $230.55) / (462+24+266 contracts) = $0.88/contract
#   mini/full: (CL $19.62 + GC $68.4 + NQ $37.8) / (6+20+12 contracts) = $3.31/contract
# Re-derived 2026-07 after Andrea flagged the reconciliation table's
# statement_fee ($787.59) vs fills_fee_reconstructed ($338.80) gap -- the
# PREVIOUS defaults here ($1.75/$6.60) were almost exactly 2x these correct
# values (ratio 1.99 on both), meaning they'd been calibrated against the
# wrong side of that reconciliation at some point. The order-history file's
# own Commission column ($0.40/contract micro, $1.00 full-size) is NOT the
# real cost -- it's only the flat exchange/clearing fee; the statement's
# real total also includes NFA/broker markup, which is why these defaults
# must come from statement_fee, not fills_fee_reconstructed.
COMMISSION_PER_MICRO = 0.88
COMMISSION_PER_MINI  = 3.31

MICRO_PREFIXES = {"MES", "MNQ", "MGC", "MCL"}
FALLBACK_CTYPE = {"ES": "MES", "NQ": "MNQ", "GC": "MGC", "CL": "CL"}

COLORS = px.colors.qualitative.Plotly + px.colors.qualitative.Dark24
THEME  = "plotly_white"


# ─────────────────────────────────────────────────────────────────────────────
# README PARSING  —  v11: returns display_name + conservative first_qty
# ─────────────────────────────────────────────────────────────────────────────

def _build_contract_label(alloc_part: str) -> str:
    """
    Convert the right-hand side of a ReadMe allocation line into a clean
    display label.

    Examples
    --------
    "3 MES"                                         → "3 MES"
    "1GC, prima erano 5 MGC …"                      → "1 GC"
    "1 MES - inattivo"                              → "1 MES  ⏸ inattivo"
    "3  MES sul secondo livello + 4 MES sul terzo"  → "3+4 MES (L2+L3)"
    """
    # Strip everything after the first comma (historical notes)
    part = alloc_part.split(",")[0].strip()

    # Normalise spacing: "1GC" → "1 GC"
    part = re.sub(r"(\d)\s*([A-Z]{2,3})", r"\1 \2", part)

    # Detect multi-level pattern  "3 MES sul secondo livello + 4 MES sul terzo livello"
    if re.search(r"sul\s+secondo\s+livello", part, re.I) and \
       re.search(r"sul\s+terzo\s+livello",   part, re.I):
        nums   = re.findall(r"(\d+)\s+([A-Z]{2,3})", part)
        ctypes = list({ct for _, ct in nums})
        ctype  = ctypes[0] if ctypes else "MES"
        qtys   = [q for q, _ in nums]
        return "+".join(qtys) + f" {ctype} (L2+L3)"

    # Single-level verbose descriptions
    part = re.sub(r"\s+sul\s+secondo\s+livello", " (L2)", part, flags=re.I)
    part = re.sub(r"\s+sul\s+terzo\s+livello",   " (L3)", part, flags=re.I)
    part = re.sub(r"\s+sul\s+primo\s+livello",   " (L1)", part, flags=re.I)

    # Inactive flag
    if re.search(r"inattiv", part, re.I):
        part = re.sub(r"\s*-?\s*inattiv\w*", "", part, flags=re.I).strip()
        part = part + "  ⏸ inattivo"

    return re.sub(r"  +", "  ", part).strip()


def parse_readme(readme_path: str) -> dict:
    """
    Parse ReadMe.txt and return a dict keyed by *normalised* name.

    Each entry::

        norm_name → {
            "display_name":   str,   # exactly as written left of ":"
            "contract_label": str,   # e.g. "3+4 MES (L2+L3)"
            "default_n":      int,   # conservative first qty (e.g. 3 for reversal)
            "total_n":        int,   # sum of all qty (e.g. 7 for reversal)
            "ctype":          str,   # primary contract type token
            "is_inactive":    bool,
        }
    """
    alloc_re = re.compile(r"(\d+)\s*([A-Z]{2,3})")
    entries  = {}
    sort_idx = 0   # preserve ReadMe line order

    try:
        with open(readme_path, "r", encoding="utf-8", errors="ignore") as fh:
            for raw in fh:
                line = raw.strip()
                if ":" not in line:
                    continue
                colon      = line.index(":")
                name_part  = line[:colon].strip()
                alloc_part = line[colon + 1:].strip()
                if not name_part or not alloc_part:
                    continue

                clean   = alloc_part.split(",")[0]
                matches = alloc_re.findall(clean)
                if not matches:
                    continue

                by_type = defaultdict(int)
                for qty_s, ctype in matches:
                    by_type[ctype] += int(qty_s)

                primary_ctype = matches[0][1]
                first_qty     = int(matches[0][0])   # conservative default
                total_qty     = by_type[primary_ctype]

                label       = _build_contract_label(alloc_part)
                is_inactive = "inattivo" in label

                norm = name_part.lower().replace("-", " ")
                norm = re.sub(r"\s+", " ", norm).strip()

                entries[norm] = {
                    "display_name":   name_part,
                    "contract_label": label,
                    "default_n":      first_qty,
                    "total_n":        total_qty,
                    "ctype":          primary_ctype,
                    "is_inactive":    is_inactive,
                    "sort_order":     sort_idx,
                }
                sort_idx += 1
    except FileNotFoundError:
        pass

    return entries


_README_PATH    = str(DATA_DIR / "ReadMe.txt")
_README_ENTRIES = parse_readme(_README_PATH)


def get_alloc(stem: str) -> dict:
    """
    Match an xlsx filename stem to a ReadMe entry via word-set containment.

    Strategy: normalise both the ReadMe entry name and the xlsx stem by
    lower-casing and replacing punctuation/underscores with spaces, then
    check whether ALL words in the ReadMe name appear in the stem's word-set.
    Pick the longest (most-specific) match to avoid ambiguity.

    This correctly handles cases like:
      ES_breakout_sessione_regolare__long  →  'ES breakout regolare - long'
      (stem has extra word 'sessione' not in ReadMe name — still matches
       because ReadMe words are a subset of stem words)

    Falls back to filename-derived name + ⚠️ if no ReadMe match found.
    """
    symbol = stem.split("_")[0].upper()

    # Normalise stem: underscores → spaces, keep all words
    norm_stem  = stem.lower().replace("__", " ").replace("_", " ")
    stem_words = set(norm_stem.split())

    best     = None
    best_len = 0
    for norm_name, entry in _README_ENTRIES.items():
        # norm_name already has dashes converted to spaces by parse_readme
        kw = set(norm_name.split())
        if kw.issubset(stem_words) and len(kw) > best_len:
            best_len = len(kw)
            best     = entry.copy()

    if best:
        ctype    = best["ctype"]
        is_micro = ctype in MICRO_PREFIXES
        return {
            "symbol":         symbol,
            "display_name":   best["display_name"],
            "contract_label": best["contract_label"],
            "default_n":      best["default_n"],
            "ctype":          ctype,
            "is_micro":       is_micro,
            "is_inactive":    best["is_inactive"],
            "matched":        True,
        }

    # No ReadMe match → show filename with warning badge; default to 0 contracts
    fallback_ctype = FALLBACK_CTYPE.get(symbol, "MES")
    fallback_name  = stem.replace("__", " ").replace("_", " ")
    return {
        "symbol":         symbol,
        "display_name":   f"⚠️ {fallback_name}",
        "contract_label": f"0 {fallback_ctype} (not in ReadMe)",
        "default_n":      0,
        "ctype":          fallback_ctype,
        "is_micro":       fallback_ctype in MICRO_PREFIXES,
        "is_inactive":    False,
        "matched":        False,
    }


# ─────────────────────────────────────────────────────────────────────────────
# DATE PARSING
# ─────────────────────────────────────────────────────────────────────────────

def parse_ts_date(val):
    """
    Parse TradeStation Italian-locale dates from openpyxl values.

    openpyxl reads DD/MM/YYYY dates where DD ≤ 12 as datetime(YYYY, DD, MM)
    because Excel auto-converted them from Italian strings to US-format dates.
    Fix: always swap month ↔ day for datetime objects coming from openpyxl.
    Strings are parsed as DD/MM/YYYY.
    """
    if val is None:
        return None
    if isinstance(val, (datetime, pd.Timestamp)):
        ts = pd.Timestamp(val)
        try:
            ts = ts.replace(month=ts.day, day=ts.month)
        except ValueError:
            pass
        return ts
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y",
                "%m/%d/%Y %H:%M", "%m/%d/%Y",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return pd.Timestamp(datetime.strptime(s, fmt))
        except ValueError:
            continue
    try:
        return pd.Timestamp(s)
    except Exception:
        return None


# ─────────────────────────────────────────────────────────────────────────────
# RAW TRADE EXTRACTION  (unchanged from v10)
# ─────────────────────────────────────────────────────────────────────────────

def extract_trades_raw(ws) -> pd.DataFrame:
    """
    Parse the 'Trades List' worksheet.

    Columns (verified across all 18 files):
      Entry row → col[0]=trade#(int), col[1]=direction, col[2]=entry_dt,
                  col[6]=n_contracts, col[7]=trade_pnl
      Exit row  → col[0]=None, col[2]=exit_dt,
                  col[6]=trade P&L ($, full position)  ← canonical
                  col[7]=cumulative P&L ($)

    P&L values already reflect TradeStation's own contract-size multiplication.
    We never multiply again.
    """
    records      = []
    header_found = False
    current      = {}

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] == "#":
                header_found = True
            continue

        if (row[0] is not None
                and isinstance(row[0], (int, float))
                and float(row[0]) == int(float(row[0]))):
            current = {
                "trade_id":    int(row[0]),
                "entry_date":  parse_ts_date(row[2]),
                "direction":   str(row[1]).strip() if row[1] else "",
                "n_contracts": int(row[6]) if row[6] is not None else 1,
            }

        elif row[0] is None and row[1] is not None and current:
            try:
                pnl     = float(row[6]) if row[6] is not None else float("nan")
                cum_pnl = float(row[7]) if row[7] is not None else float("nan")
            except (TypeError, ValueError):
                current = {}
                continue

            records.append({
                "trade_id":    current["trade_id"],
                "entry_date":  current["entry_date"],
                "exit_date":   parse_ts_date(row[2]),
                "direction":   current["direction"],
                "n_contracts": current["n_contracts"],
                "pnl":         pnl,
                "cum_pnl":     cum_pnl,
            })
            current = {}

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)
    df["pnl"]     = pd.to_numeric(df["pnl"],     errors="coerce")
    df["cum_pnl"] = pd.to_numeric(df["cum_pnl"], errors="coerce")
    return (df.dropna(subset=["exit_date", "pnl"])
              .sort_values("exit_date")
              .reset_index(drop=True))


def _snap_to_bday(dt: pd.Timestamp) -> pd.Timestamp:
    """Snap weekend dates to next Monday (handles GC weekend-bias exits)."""
    return dt if dt.dayofweek < 5 else dt + pd.offsets.BDay(1)


def build_equity_from_trades(trades: pd.DataFrame) -> pd.Series:
    """
    Build daily equity curve (cumsum of daily P&L) from raw trades.
    Weekend exit dates are snapped to the next business day.
    Forward-fills over business-day calendar for continuity.
    """
    if trades.empty:
        return pd.Series(dtype=float)
    t = trades.copy()
    t["exit_day"] = pd.to_datetime(t["exit_date"]).dt.normalize().apply(_snap_to_bday)
    daily  = t.groupby("exit_day")["pnl"].sum().sort_index()
    equity = daily.cumsum()
    if len(equity) < 2:
        return equity
    all_days = pd.date_range(equity.index.min(), equity.index.max(), freq="B")
    return equity.reindex(all_days).ffill()


# ─────────────────────────────────────────────────────────────────────────────
# LOAD ALL SYSTEMS
# ─────────────────────────────────────────────────────────────────────────────

_LOAD_VERSION = 6   # bump this integer to force a full cache clear
                    # (bumped 5→6: adds .csv strategy-file support below)

@st.cache_data(show_spinner=False, ttl=86400)
def load_all_systems(data_dir: str, comm_micro: float, comm_mini: float,
                     _version: int = _LOAD_VERSION):
    """
    Cached with st.cache_data (TTL 24 h).
    Changing _LOAD_VERSION forces a fresh load.
    Sizing lives in session_state and is applied as a ratio at render time —
    so slider changes never trigger xlsx re-reads.

    Supports both .xlsx (openpyxl "Trades List" sheet) and .csv (flat report
    export — used when TradeStation's xlsx report generator is unavailable/
    crashing) per system, same convention as the Trade Matching tab: if a
    system has BOTH an .xlsx and a .csv with the same filename stem, the
    .csv is used and the .xlsx is skipped, since .csv is the newer/likely
    more current export path for this account.
    """
    path          = Path(data_dir)
    systems       = {}
    load_warnings = []

    xlsx_files = {f.stem: f for f in sorted(path.glob("*.xlsx"))}
    csv_files  = {f.stem: f for f in sorted(path.glob("*.csv"))}
    stems      = sorted(set(xlsx_files) | set(csv_files))

    for stem in stems:
        f = csv_files.get(stem) or xlsx_files[stem]
        from_csv = stem in csv_files
        try:
            b5_val      = None
            data_source = "trades"
            trades      = pd.DataFrame()

            if from_csv:
                if tsm is None:
                    load_warnings.append(
                        f"❌ {f.name}: found a .csv strategy file but "
                        f"tradestation_matcher.py couldn't be imported "
                        f"({_TSM_IMPORT_ERROR}) — skipped.")
                    continue
                trades = tsm.extract_theoretical_trades_csv(str(f))
                b5_val = tsm.extract_csv_total_net_profit(str(f))
                if trades.empty:
                    load_warnings.append(f"⚠️ {f.name}: 'Trades List' section is empty.")
                    data_source = "empty"
            else:
                wb = openpyxl.load_workbook(str(f), data_only=True)

                if "Performance Summary" in wb.sheetnames:
                    try:
                        b5_val = float(wb["Performance Summary"]["B5"].value)
                    except (TypeError, ValueError):
                        pass

                if "Trades List" in wb.sheetnames:
                    trades = extract_trades_raw(wb["Trades List"])
                    if trades.empty:
                        load_warnings.append(f"⚠️ {f.name}: 'Trades List' is empty.")
                        data_source = "empty"
                else:
                    load_warnings.append(
                        f"⚠️ {f.name}: no 'Trades List' tab — using summary only.")
                    data_source = "summary_fallback"

            equity = build_equity_from_trades(trades)

            b5_ok = False
            if not trades.empty and b5_val is not None:
                last_cum = trades["cum_pnl"].iloc[-1]
                b5_ok = abs(float(b5_val) - float(last_cum)) < 0.5
                if not b5_ok:
                    load_warnings.append(
                        f"⚠️ {stem}: last cum_pnl ≠ B5 ({last_cum} vs {b5_val})")

            alloc    = get_alloc(stem)
            ctype    = alloc["ctype"]
            is_micro = ctype in MICRO_PREFIXES
            comm_rt  = comm_micro if is_micro else comm_mini

            systems[stem] = {
                # ── Identity ──────────────────────────────────────────────
                "stem":           stem,
                "display_name":   alloc["display_name"],   # from ReadMe
                "symbol":         alloc["symbol"],
                "contract_type":  ctype,
                "contract_label": alloc["contract_label"],
                "default_n":      alloc["default_n"],       # ReadMe first qty
                "is_micro":       is_micro,
                "is_inactive":    alloc["is_inactive"],
                "matched":        alloc["matched"],
                # ── Trade data ────────────────────────────────────────────
                "trades":         trades,
                "equity":         equity,               # raw, 1× factor
                # ── Costs ─────────────────────────────────────────────────
                "comm_per_trade": comm_rt,
                # ── Validation ────────────────────────────────────────────
                "b5_total":       b5_val,
                "b5_match":       b5_ok,
                "data_source":    data_source,
                "from_csv":       from_csv,
            }
        except Exception as exc:
            load_warnings.append(f"❌ Could not load {f.name}: {exc}")

    # ── Sort by ReadMe.txt line order (sort_order baked into each entry) ──────
    def _sort_key(stem):
        norm_stem  = stem.lower().replace("__", " ").replace("_", " ")
        stem_words = set(norm_stem.split())
        best_order = 9999
        best_len   = 0
        for norm_name, entry in _README_ENTRIES.items():
            kw = set(norm_name.split())
            if kw.issubset(stem_words) and len(kw) > best_len:
                best_len   = len(kw)
                best_order = entry.get("sort_order", 9999)
        return best_order

    systems = dict(sorted(systems.items(), key=lambda kv: _sort_key(kv[0])))

    return systems, load_warnings


# ─────────────────────────────────────────────────────────────────────────────
# SIZING HELPERS  (v11 — session_state sizing multiplier)
# ─────────────────────────────────────────────────────────────────────────────

def get_current_n(stem: str, systems: dict) -> int:
    """Return current user-set contract count from session_state, or default."""
    return st.session_state.get("sizing", {}).get(
        stem, systems[stem]["default_n"])


def get_sizing_ratio(stem: str, systems: dict) -> float:
    """
    Return scaling ratio = current_n / default_n.

    This is the ONLY place a multiplier is applied to raw P&L.
    It scales the displayed equity curve to reflect a different
    number of contracts than what TradeStation ran.

    If default_n == 0, return 0 to avoid division by zero.
    """
    default = systems[stem]["default_n"]
    if default == 0:
        return 0.0
    return get_current_n(stem, systems) / default


# ─────────────────────────────────────────────────────────────────────────────
# METRICS
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False, ttl=3600)
def build_net_equity(trades: pd.DataFrame, comm_per_trade: float,
                     scale: float = 1.0) -> pd.Series:
    """
    Build daily equity after subtracting commission.
    `comm_per_trade` is a $/CONTRACT round-trip rate (not a flat per-trade
    cost) -- multiplied by each trade's actual n_contracts below. Confirmed
    against Andrea's real statement: commissions are charged per contract,
    so a flat per-trade deduction badly underestimated cost on multi-lot
    trades (and overestimated it on the single-lot full-size systems).
    'scale' is the sizing ratio (current_n / default_n), applied
    multiplicatively to the net P&L series on top of that.
    Cached: re-computes only when trades, comm or scale actually change.
    """
    if trades.empty:
        return pd.Series(dtype=float)
    t = trades.copy()
    t["exit_day"] = pd.to_datetime(t["exit_date"]).dt.normalize().apply(_snap_to_bday)
    n_ct = t["n_contracts"] if "n_contracts" in t.columns else 1
    t["pnl_net"]  = (t["pnl"] - comm_per_trade * n_ct) * scale
    daily = t.groupby("exit_day")["pnl_net"].sum().sort_index()
    eq    = daily.cumsum()
    if len(eq) < 2:
        return eq
    all_days = pd.date_range(eq.index.min(), eq.index.max(), freq="B")
    return eq.reindex(all_days).ffill()


def compute_metrics(trades: pd.DataFrame, comm_per_trade: float,
                    scale: float = 1.0) -> dict:
    """
    Compute all performance metrics.
    'scale' = current_n / default_n (1.0 = ReadMe default).
    """
    if trades.empty:
        return {}

    eq_net = build_net_equity(trades, comm_per_trade, scale)
    if eq_net.empty or len(eq_net) < 5:
        return {}

    daily   = eq_net.diff().dropna()
    run_max = eq_net.cummax()
    max_dd  = (eq_net - run_max).min()
    n_yr    = max((eq_net.index[-1] - eq_net.index[0]).days / 365.25, 0.01)
    net     = eq_net.iloc[-1] - eq_net.iloc[0]
    ann_r   = net / n_yr
    ann_v   = daily.std() * np.sqrt(252)
    sharpe  = ann_r / ann_v       if ann_v  > 0 else 0.0
    calmar  = ann_r / abs(max_dd) if max_dd < 0 else 0.0

    n_ct_t = trades["n_contracts"] if "n_contracts" in trades.columns else 1
    pnl_net_t = (trades["pnl"] - comm_per_trade * n_ct_t) * scale
    wins_sum   = pnl_net_t[pnl_net_t > 0].sum()
    loss_sum   = abs(pnl_net_t[pnl_net_t < 0].sum())
    pf         = wins_sum / loss_sum if loss_sum > 0 else float("inf")
    win_rate   = (pnl_net_t > 0).mean()

    return {
        "Net Profit ($)":   round(net, 0),
        "Ann. Return ($)":  round(ann_r, 0),
        "Max Drawdown ($)": round(max_dd, 0),
        "Sharpe Ratio":     round(sharpe, 2),
        "Calmar Ratio":     round(calmar, 2),
        "Profit Factor":    round(pf, 2),
        "Win Rate":         round(win_rate, 3),
        "Ann. Volatility":  round(ann_v, 0),
        "# Trades":         len(trades),
        "Total Comm ($)":   round(float((n_ct_t * comm_per_trade).sum()) * scale, 0),
    }


def get_net_equity_trimmed(si: dict, lookback_years: float,
                            scale: float = 1.0) -> pd.Series:
    """Return lookback-clipped, rebased-to-zero net equity for a system."""
    eq = build_net_equity(si["trades"], si["comm_per_trade"], scale)
    if eq.empty:
        return pd.Series(dtype=float)
    cutoff = eq.index.max() - pd.Timedelta(days=int(lookback_years * 365))
    eq_t   = eq[eq.index >= cutoff]
    if eq_t.empty:
        return pd.Series(dtype=float)
    return eq_t - eq_t.iloc[0]


def compute_portfolio_metrics(port_eq: pd.Series) -> dict:
    if port_eq.empty or len(port_eq) < 5:
        return {}
    daily  = port_eq.diff().dropna()
    net    = port_eq.iloc[-1] - port_eq.iloc[0]
    n_yr   = max((port_eq.index[-1] - port_eq.index[0]).days / 365.25, 0.01)
    ann_r  = net / n_yr
    ann_v  = daily.std() * np.sqrt(252)
    sharpe = ann_r / ann_v if ann_v > 0 else 0.0
    dd     = (port_eq - port_eq.cummax()).min()
    calmar = ann_r / abs(dd) if dd < 0 else 0.0
    return {
        "Net Profit ($)":   round(net, 0),
        "Ann. Return ($)":  round(ann_r, 0),
        "Max Drawdown ($)": round(dd, 0),
        "Ann. Volatility":  round(ann_v, 0),
        "Sharpe Ratio":     round(sharpe, 2),
        "Calmar Ratio":     round(calmar, 2),
    }


def combine_equity_curves(curves: dict, lookback_years: float) -> pd.DataFrame:
    if not curves:
        return pd.DataFrame()
    df     = pd.concat(curves.values(), axis=1, keys=curves.keys()).sort_index().ffill()
    cutoff = df.index.max() - pd.Timedelta(days=int(lookback_years * 365))
    return df[df.index >= cutoff].ffill().fillna(0)


# ─────────────────────────────────────────────────────────────────────────────
# ROLLING HEALTH
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False, ttl=3600)
def compute_rolling_health(trades: pd.DataFrame, comm_per_trade: float,
                            window_days: int = 63) -> pd.DataFrame:
    eq_net = build_net_equity(trades, comm_per_trade)
    if eq_net.empty or len(eq_net) < window_days + 10:
        return pd.DataFrame()
    daily = eq_net.diff().dropna()
    if len(daily) < window_days:
        return pd.DataFrame()

    roll_sharpe = (daily.rolling(window_days).mean()
                   / daily.rolling(window_days).std()
                   * np.sqrt(252)).dropna()
    wins_r = (daily > 0).astype(float).rolling(window_days).mean().dropna()
    gains  = daily.clip(lower=0).rolling(window_days).sum()
    losses = (-daily).clip(lower=0).rolling(window_days).sum()
    roll_pf  = (gains / losses.replace(0, np.nan)).dropna()
    roll_cum = daily.rolling(window_days).sum().dropna()

    common = roll_sharpe.index.intersection(wins_r.index)
    if common.empty:
        return pd.DataFrame()

    return pd.DataFrame({
        "Sharpe":        roll_sharpe.reindex(common),
        "Win Rate":      wins_r.reindex(common),
        "Profit Factor": roll_pf.reindex(common) if not roll_pf.empty else np.nan,
        "Period P&L":    roll_cum.reindex(common),
    }).dropna()


def health_traffic_light(sharpe: float, pf: float, win_rate: float) -> str:
    score  = 0
    score += 2 if sharpe   > 0.8  else (1 if sharpe   > 0.2  else (-1 if sharpe   < -0.3 else 0))
    score += 2 if pf       > 1.3  else (1 if pf       > 1.0  else (-1 if pf       < 0.8  else 0))
    score += 1 if win_rate > 0.55 else (-1 if win_rate < 0.4 else 0)
    return "🟢" if score >= 3 else ("🟡" if score >= 1 else "🔴")


# ─────────────────────────────────────────────────────────────────────────────
# PORTFOLIO HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def build_portfolio_equity(systems: dict, stems: list,
                            lookback_years: float,
                            sizing_ratios: dict | None = None) -> pd.Series:
    """
    Sum per-system net-equity curves.
    sizing_ratios = {stem: float}, defaults to 1.0 (ReadMe default).
    """
    curves = {}
    for stem in stems:
        si    = systems.get(stem)
        if si is None:
            continue
        ratio = (sizing_ratios or {}).get(stem, 1.0)
        if ratio == 0.0:
            continue   # system switched off — skip entirely
        eq    = get_net_equity_trimmed(si, lookback_years, scale=ratio)
        if not eq.empty:
            curves[stem] = eq
    if not curves:
        return pd.Series(dtype=float)
    df = pd.concat(curves.values(), axis=1,
                   keys=curves.keys()).sort_index().ffill().fillna(0)
    return df.sum(axis=1)


# ─────────────────────────────────────────────────────────────────────────────
# RISK PARITY
# ─────────────────────────────────────────────────────────────────────────────

def compute_risk_parity_sizing(systems: dict, stems: list, lookback_years: float,
                                target_daily_risk_usd: float = 500.0,
                                max_contracts: int = 20) -> tuple:
    sizing, detail = {}, {}
    for stem in stems:
        si = systems.get(stem)
        if si is None or si["equity"].empty:
            sizing[stem] = 1
            continue
        eq     = si["equity"]
        cutoff = eq.index.max() - pd.Timedelta(days=int(lookback_years * 365))
        eq_t   = eq[eq.index >= cutoff]
        vol_1  = eq_t.diff().dropna().std()
        if vol_1 <= 0 or np.isnan(vol_1):
            sizing[stem] = si["default_n"]
            detail[stem] = {"vol_1ct": 0, "raw_n": si["default_n"], "final_n": si["default_n"]}
            continue
        raw_n   = target_daily_risk_usd / vol_1
        final_n = int(np.clip(round(raw_n), 0, max_contracts))
        sizing[stem] = final_n
        detail[stem] = {"vol_1ct": round(float(vol_1), 2),
                        "raw_n":   round(float(raw_n), 2),
                        "final_n": final_n}
    return sizing, detail


# ─────────────────────────────────────────────────────────────────────────────
# DRAWDOWN DECOMPOSITION
# ─────────────────────────────────────────────────────────────────────────────

def decompose_drawdown(eq_df: pd.DataFrame, port_eq: pd.Series, top_n: int = 3) -> list:
    if port_eq.empty or eq_df.empty:
        return []
    peak_idx, peak_val = 0, port_eq.iloc[0]
    episodes = []
    for i in range(1, len(port_eq)):
        val = port_eq.iloc[i]
        if val > peak_val:
            peak_val, peak_idx = val, i
        else:
            dd = val - peak_val
            if not episodes or episodes[-1]["peak_idx"] != peak_idx:
                episodes.append({"peak_idx": peak_idx, "trough_idx": i,
                                  "dd_abs": dd,
                                  "peak_date":   port_eq.index[peak_idx],
                                  "trough_date": port_eq.index[i]})
            elif dd < episodes[-1]["dd_abs"]:
                episodes[-1].update({"trough_idx": i, "dd_abs": dd,
                                      "trough_date": port_eq.index[i]})
    episodes = sorted(episodes, key=lambda x: x["dd_abs"])[:top_n]
    for ep in episodes:
        pd_, td = ep["peak_date"], ep["trough_date"]
        contribs = {}
        for col in eq_df.columns:
            s = eq_df[col]
            try:
                s_p = float(s.asof(pd_)) if pd_ >= s.index[0] else 0.0
                s_t = float(s.asof(td))  if td  >= s.index[0] else 0.0
                contribs[col] = s_t - s_p
            except Exception:
                contribs[col] = 0.0
        total_loss = abs(ep["dd_abs"])
        ep["contributions"] = contribs
        ep["blame_pct"]     = {k: abs(v) / total_loss * 100
                               for k, v in contribs.items()
                               if v < 0 and total_loss > 0}
    return episodes


# ─────────────────────────────────────────────────────────────────────────────
# CORRELATION / CLUSTERING
# ─────────────────────────────────────────────────────────────────────────────

def cluster_correlation_matrix(corr_matrix: pd.DataFrame) -> tuple:
    # Drop any rows/cols that are all-NaN (systems with no data / n=0)
    corr_matrix = corr_matrix.dropna(axis=0, how="all").dropna(axis=1, how="all")
    n = len(corr_matrix)
    if n < 3:
        return corr_matrix, list(range(n))
    # Ensure symmetry: average with transpose to fix floating-point asymmetry
    vals = corr_matrix.values.clip(-1, 1)
    vals = (vals + vals.T) / 2
    np.fill_diagonal(vals, 1.0)
    dist = np.maximum(1 - vals, 0)
    # squareform expects a condensed vector; pass the full symmetric matrix
    try:
        from scipy.spatial.distance import squareform as sf
        dist_condensed = sf(dist, checks=False)
        order = leaves_list(linkage(dist_condensed, method="ward"))
    except Exception:
        order = list(range(n))
    cols = corr_matrix.columns[order]
    return corr_matrix.loc[cols, cols], order


def compute_cluster_risk_score(corr_matrix: pd.DataFrame) -> float:
    if corr_matrix.empty or len(corr_matrix) < 2:
        return 0.0
    upper = corr_matrix.where(np.triu(np.ones(corr_matrix.shape), k=1).astype(bool))
    vals  = upper.stack()
    return float(np.clip((vals.mean() + 1) / 2 * 100, 0, 100)) if not vals.empty else 0.0


# ─────────────────────────────────────────────────────────────────────────────
# PORTFOLIO OPTIMISATION HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def recommend_portfolios(eq_df: pd.DataFrame, systems: dict) -> list:
    daily = eq_df.diff().dropna()
    if daily.empty or daily.shape[1] < 2:
        return []
    means   = daily.mean()
    stds    = daily.std().replace(0, np.nan)
    sharpes = (means / stds).fillna(0)
    corr    = daily.corr()
    recs    = []

    top = sharpes.nlargest(min(8, len(sharpes))).index.tolist()
    sel = []
    for s in top:
        if not sel or max(corr.loc[s, x] for x in sel if x in corr.columns) < 0.85:
            sel.append(s)
    recs.append({"name": "🏆 Max Sharpe",
                 "description": "Top-Sharpe, pairwise ρ < 0.85", "systems": sel})

    n    = min(6, len(corr))
    seed = sharpes.idxmax()
    grp  = [seed]
    rem  = [x for x in corr.columns if x != seed]
    while len(grp) < n and rem:
        avg_c = {r: corr.loc[r, grp].mean() for r in rem}
        nxt   = min(avg_c, key=avg_c.get)
        grp.append(nxt); rem.remove(nxt)
    recs.append({"name": "🌐 Max Diversification",
                 "description": "Min avg pairwise correlation", "systems": grp})

    scores = {}
    for col in daily.columns:
        cum = daily[col].cumsum()
        dd  = (cum - cum.cummax()).min()
        scores[col] = means[col] / abs(dd) if dd < 0 else 0.0
    recs.append({"name": "🛡 Min Drawdown",
                 "description": "Best return/drawdown ratio",
                 "systems": sorted(scores, key=scores.get, reverse=True)[:6]})
    return recs


# ─────────────────────────────────────────────────────────────────────────────
# MONTE CARLO
# ─────────────────────────────────────────────────────────────────────────────

def monte_carlo_simulation(systems: dict, stems: list, lookback_years: float,
                            n_sims: int = 1000, forward_days: int = 126,
                            sizing_ratios: dict | None = None) -> dict | None:
    daily_list = []
    for stem in stems:
        si    = systems.get(stem)
        if si is None or si["trades"].empty:
            continue
        ratio = (sizing_ratios or {}).get(stem, 1.0)
        eq    = get_net_equity_trimmed(si, lookback_years, scale=ratio)
        if not eq.empty:
            daily_list.append(eq.diff().dropna())
    if not daily_list:
        return None
    pool = pd.concat(daily_list, axis=1).fillna(0).sum(axis=1).values
    if len(pool) < 20:
        return None
    rng   = np.random.default_rng(42)
    paths = np.zeros((forward_days, n_sims))
    for sim in range(n_sims):
        paths[:, sim] = np.cumsum(rng.choice(pool, size=forward_days, replace=True))
    pctiles = pd.DataFrame({
        k: np.percentile(paths, p, axis=1)
        for k, p in [("5th", 5), ("25th", 25), ("50th", 50), ("75th", 75), ("95th", 95)]
    })
    peak  = np.maximum.accumulate(paths, axis=0)
    dd_probs = {t: ((paths - peak).min(axis=0) < -t).mean()
                for t in [5000, 10000, 15000, 20000, 30000, 50000]}
    return {"paths": pd.DataFrame(paths), "percentiles": pctiles,
            "dd_probs": dd_probs, "n_sims": n_sims, "forward_days": forward_days}


# ─────────────────────────────────────────────────────────────────────────────
# REGIME ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────

def compute_regime_series(equity_dict: dict, vol_window: int = 20,
                           low_pct: int = 33, high_pct: int = 66) -> tuple:
    if not equity_dict:
        return pd.DataFrame(), {}
    df    = pd.concat(equity_dict.values(), axis=1,
                      keys=equity_dict.keys()).sort_index().ffill().fillna(0)
    daily = df.sum(axis=1).diff().dropna()
    if len(daily) < vol_window + 10:
        return pd.DataFrame(), {}
    roll_vol = (daily.rolling(vol_window).std() * np.sqrt(252)).dropna()
    if roll_vol.empty:
        return pd.DataFrame(), {}
    lo, hi = np.percentile(roll_vol, low_pct), np.percentile(roll_vol, high_pct)
    def _cls(v):
        return "Low Vol" if v <= lo else ("High Vol" if v >= hi else "Medium Vol")
    return (pd.DataFrame({"vol": roll_vol, "regime": roll_vol.apply(_cls)}),
            {"low": lo, "high": hi})


def compute_system_regime_performance(trades: pd.DataFrame, regime_df: pd.DataFrame,
                                       comm_per_trade: float) -> dict:
    if trades.empty or regime_df.empty:
        return {}
    t = trades.copy()
    t["exit_day"] = pd.to_datetime(t["exit_date"]).dt.normalize()
    n_ct_r = t["n_contracts"] if "n_contracts" in t.columns else 1
    t["pnl_net"]  = t["pnl"] - comm_per_trade * n_ct_r

    def _nr(dt):
        if dt in regime_df.index:
            return regime_df.loc[dt, "regime"]
        for off in range(1, 6):
            prev = dt - pd.Timedelta(days=off)
            if prev in regime_df.index:
                return regime_df.loc[prev, "regime"]
        return None

    t["regime"] = t["exit_day"].apply(_nr)
    t = t.dropna(subset=["regime"])
    if t.empty:
        return {}

    results = {}
    for lbl in ["Low Vol", "Medium Vol", "High Vol"]:
        g = t[t["regime"] == lbl]
        if g.empty:
            results[lbl] = {"n_trades": 0, "total_pnl": 0, "avg_pnl": 0,
                             "win_rate": 0, "profit_factor": 0}
            continue
        wins  = g["pnl_net"] > 0
        gains = g.loc[wins, "pnl_net"].sum()
        loss  = abs(g.loc[~wins, "pnl_net"].sum())
        results[lbl] = {
            "n_trades":      len(g),
            "total_pnl":     round(g["pnl_net"].sum(), 0),
            "avg_pnl":       round(g["pnl_net"].mean(), 0),
            "win_rate":      round(wins.mean(), 3),
            "profit_factor": round(gains / loss if loss > 0 else (np.inf if gains > 0 else 0), 2),
        }
    return results


def compute_regime_transitions(regime_df: pd.DataFrame) -> pd.DataFrame:
    if regime_df.empty or len(regime_df) < 2:
        return pd.DataFrame()
    labels = ["Low Vol", "Medium Vol", "High Vol"]
    trans  = pd.DataFrame(0, index=labels, columns=labels, dtype=float)
    counts = pd.Series(0, index=labels, dtype=float)
    regs   = regime_df["regime"]
    for i in range(len(regs) - 1):
        c, n = regs.iloc[i], regs.iloc[i + 1]
        if c in labels and n in labels:
            trans.loc[c, n] += 1
            counts[c]       += 1
    for r in labels:
        if counts[r] > 0:
            trans.loc[r] = trans.loc[r] / counts[r]
    return trans.round(3)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def build_excel_export(systems: dict, lookback_years: float,
                        sizing_ratios: dict | None = None,
                        corr_matrix=None) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR_FILL = PatternFill("solid", fgColor="1F3864")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    ALT_FILL = PatternFill("solid", fgColor="F2F4F8")

    def _ws(name, headers, rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for c in ws[1]:
            c.fill, c.font = HDR_FILL, HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
        for i, row in enumerate(rows, 2):
            ws.append(row)
            if i % 2 == 0:
                for c in ws[i]:
                    c.fill = ALT_FILL
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 40)
        ws.freeze_panes = "A2"
        return ws

    ratios = sizing_ratios or {}

    # Summary sheet
    hdr1  = ["System", "Contract Label", "Default N", "Current N", "Scale",
             "Source", "# Trades", "Net Profit ($)", "Max DD ($)",
             "Sharpe", "Calmar", "PF", "Win Rate", "B5 Match"]
    rows1 = []
    for stem, si in systems.items():
        ratio = ratios.get(stem, 1.0)
        cur_n = round(si["default_n"] * ratio)
        m     = compute_metrics(si["trades"], si["comm_per_trade"], ratio)
        rows1.append([
            si["display_name"][:40], si["contract_label"],
            si["default_n"], cur_n, f"{ratio:.2f}×",
            "📊 trades" if si["data_source"] == "trades" else "⚠️ fallback",
            m.get("# Trades", 0),
            m.get("Net Profit ($)", ""),    m.get("Max Drawdown ($)", ""),
            m.get("Sharpe Ratio", ""),      m.get("Calmar Ratio", ""),
            m.get("Profit Factor", ""),
            f'{m.get("Win Rate", 0):.1%}' if m else "",
            "✅" if si["b5_match"] else "❌",
        ])
    _ws("Summary", hdr1, rows1)

    # Trade log
    hdr2  = ["System", "Trade#", "Direction", "Entry Date", "Exit Date",
             "N Contracts (TS)", "P&L ($) Raw", "Cum P&L ($)"]
    rows2 = []
    for stem, si in systems.items():
        for _, r in si["trades"].iterrows():
            rows2.append([
                si["display_name"][:30], r["trade_id"], r["direction"],
                str(r["entry_date"])[:16] if pd.notna(r.get("entry_date")) else "",
                str(r["exit_date"])[:16],
                r.get("n_contracts", 1),
                round(r["pnl"], 2), round(r["cum_pnl"], 2),
            ])
    _ws("Trade Log", hdr2, rows2)

    # Correlation
    if corr_matrix is not None and not corr_matrix.empty:
        ws3  = wb.create_sheet("Correlation")
        cols = [systems[c]["display_name"][:20] if c in systems else c[:20]
                for c in corr_matrix.columns]
        ws3.append([""] + cols)
        for idx, row_data in corr_matrix.iterrows():
            label = systems[idx]["display_name"][:20] if idx in systems else idx[:20]
            ws3.append([label] + [round(v, 3) for v in row_data.values])
        ws3.column_dimensions["A"].width = 28
        lc = get_column_letter(len(cols) + 1)
        lr = len(cols) + 1
        ws3.conditional_formatting.add(
            f"B2:{lc}{lr}",
            ColorScaleRule(start_type="num", start_value=-1, start_color="D73027",
                           mid_type="num",   mid_value=0,    mid_color="FFFFFF",
                           end_type="num",   end_value=1,    end_color="4575B4"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# PLOTTING
# ─────────────────────────────────────────────────────────────────────────────

C = {
    "drawdown":  "rgba(220,50,47,0.6)",
    "dd_line":   "#dc322f",
    "peak":      "#93a1a1",
    "portfolio": "#073642",
    "green":     "#2aa198",
    "red":       "#dc322f",
    "amber":     "#b58900",
    "blue":      "#268bd2",
    "zero_line": "#93a1a1",
}


def plot_equity(eq: pd.Series, name: str, color: str = "#268bd2") -> go.Figure:
    run_max = eq.cummax()
    dd      = eq - run_max
    fig     = make_subplots(rows=2, cols=1, shared_xaxes=True,
                            row_heights=[0.7, 0.3], vertical_spacing=0.04)
    fig.add_trace(go.Scatter(x=eq.index, y=eq.values, name=name,
                             line=dict(color=color, width=2),
                             hovertemplate="%{x|%Y-%m-%d}<br>$%{y:,.0f}<extra></extra>"),
                  row=1, col=1)
    fig.add_trace(go.Scatter(x=eq.index, y=run_max.values, name="Peak",
                             line=dict(color=C["peak"], width=1, dash="dot"),
                             showlegend=False), row=1, col=1)
    fig.add_trace(go.Scatter(x=eq.index, y=dd.values, name="DD",
                             fill="tozeroy", line=dict(color=C["dd_line"], width=1),
                             fillcolor=C["drawdown"],
                             hovertemplate="%{x|%Y-%m-%d}<br>$%{y:,.0f}<extra></extra>"),
                  row=2, col=1)
    fig.update_layout(template=THEME, height=500,
                      margin=dict(l=0, r=0, t=10, b=0),
                      legend=dict(orientation="h", y=1.05),
                      yaxis_title="Cum. P&L ($)", yaxis2_title="Drawdown ($)")
    return fig


def plot_portfolio_equity(port_eq: pd.Series, eq_df: pd.DataFrame,
                           label_map: dict | None = None) -> go.Figure:
    fig = make_subplots(rows=2, cols=1, shared_xaxes=True,
                        row_heights=[0.65, 0.35], vertical_spacing=0.04)
    for i, col in enumerate(eq_df.columns):
        rc     = COLORS[i % len(COLORS)]
        fill_c = rc.replace("rgb", "rgba").replace(")", ",0.45)") if rc.startswith("rgb") else rc
        lbl    = (label_map or {}).get(col, col.replace("_", " ")[:30])
        fig.add_trace(go.Scatter(
            x=eq_df.index, y=eq_df[col].values,
            name=lbl, stackgroup="one",
            line=dict(width=0.5), fillcolor=fill_c,
            hovertemplate=f"{lbl[:25]}<br>%{{x|%Y-%m-%d}}<br>$%{{y:,.0f}}<extra></extra>"),
            row=1, col=1)
    fig.add_trace(go.Scatter(x=port_eq.index, y=port_eq.values,
                             name="Portfolio Total",
                             line=dict(color=C["portfolio"], width=2.5)),
                  row=1, col=1)
    dd = port_eq - port_eq.cummax()
    fig.add_trace(go.Scatter(x=port_eq.index, y=dd.values, name="DD",
                             fill="tozeroy", line=dict(color=C["dd_line"], width=1),
                             fillcolor=C["drawdown"]), row=2, col=1)
    fig.update_layout(template=THEME, height=580,
                      margin=dict(l=0, r=0, t=10, b=0),
                      yaxis_title="Cum. P&L ($)", yaxis2_title="Drawdown ($)")
    return fig


def plot_monthly_heatmap(equity: pd.Series, name: str) -> go.Figure:
    """
    Monthly P&L heatmap with an annual 'Total' column.

    The 12 monthly cells use RdYlGn coloring relative to their own range.
    The 'Total' column uses a *separate* heatmap trace with independent
    coloring so year totals don't distort the monthly color scale and vice versa.
    A thin vertical gap visually separates the two sections.
    """
    monthly = equity.resample("ME").last().diff().dropna()
    if monthly.empty:
        return go.Figure()

    df = pd.DataFrame({"val": monthly})
    df["year"]  = df.index.year
    df["month"] = df.index.month
    pivot = df.pivot(index="year", columns="month", values="val").fillna(0)

    mnames = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    # Only keep columns that exist in the data
    month_labels = [mnames[c-1] for c in pivot.columns]
    years        = [str(y) for y in pivot.index.tolist()]

    monthly_z    = pivot.values.tolist()
    monthly_text = [[f"${v:,.0f}" for v in r] for r in pivot.values]

    # Annual totals
    totals      = pivot.sum(axis=1).values.tolist()
    totals_text = [[f"${v:,.0f}"] for v in totals]
    totals_z    = [[v] for v in totals]

    # x-positions: months at 0..N-1, gap, Total at N+0.6
    n_months  = len(month_labels)
    x_months  = list(range(n_months))
    x_total   = [n_months + 0.6]
    x_labels  = month_labels + [""] + ["Total"]   # blank spacer tick

    fig = go.Figure()

    # ── Monthly heatmap ───────────────────────────────────────────────────────
    # Colorbars are placed AFTER the Total column so they never overlap the
    # month cells. Monthly bar sits just right of Total; Annual bar further right.
    # thickness=12 keeps them slim; xpad=4 adds a small gap.
    fig.add_trace(go.Heatmap(
        z=monthly_z,
        x=x_months,
        y=years,
        text=monthly_text,
        texttemplate="%{text}",
        colorscale="RdYlGn",
        zmid=0,
        showscale=True,
        colorbar=dict(
            title=dict(text="Monthly", side="right"),
            x=1.06,           # well right of the Total column
            xpad=4,
            thickness=12,
            len=0.85,
        ),
        hovertemplate="Year %{y}<br>%{x}<br>$%{z:,.0f}<extra></extra>",
        name="Monthly",
    ))

    # ── Annual total heatmap (independent color scale) ────────────────────────
    fig.add_trace(go.Heatmap(
        z=totals_z,
        x=x_total,
        y=years,
        text=totals_text,
        texttemplate="%{text}",
        colorscale="RdYlGn",
        zmid=0,
        showscale=True,
        colorbar=dict(
            title=dict(text="Annual", side="right"),
            x=1.18,           # further right, no overlap with monthly bar
            xpad=4,
            thickness=12,
            len=0.85,
        ),
        hovertemplate="Year %{y}<br>Total: $%{z:,.0f}<extra></extra>",
        name="Annual Total",
    ))

    # Build tick labels: month names at integer positions, blank spacer, "Total"
    tickvals = x_months + [n_months + 0.0, n_months + 0.6]
    ticktext = month_labels + ["", "Total"]

    fig.update_layout(
        template=THEME,
        height=max(220, 34 * len(pivot) + 80),
        # right margin large enough to show both slim colorbars without overlap
        margin=dict(l=0, r=160, t=40, b=0),
        title=f"Monthly P&L — {name}",
        xaxis=dict(
            tickmode="array",
            tickvals=tickvals,
            ticktext=ticktext,
            tickangle=0,
        ),
    )
    return fig


def plot_clustered_correlation(corr_matrix: pd.DataFrame, systems: dict,
                                over_thresh: float = 0.70) -> go.Figure:
    reordered, _ = cluster_correlation_matrix(corr_matrix)
    cols   = reordered.columns.tolist()
    labels = [systems[c]["display_name"][:22] if c in systems else c[:22] for c in cols]
    shapes = []
    n      = len(cols)
    for i in range(n):
        for j in range(i+1, n):
            if reordered.iloc[i, j] > over_thresh:
                shapes.append(dict(type="rect",
                    x0=j-0.5, x1=j+0.5, y0=i-0.5, y1=i+0.5,
                    line=dict(color="rgba(220,50,47,0.9)", width=2)))
    colorscale = [[0.0,"#d73027"],[0.4,"#f7b99e"],
                  [0.5,"#ffffff"],[0.6,"#92c5de"],[1.0,"#4575b4"]]
    fig = go.Figure(go.Heatmap(
        z=reordered.values, x=labels, y=labels,
        text=[[f"{v:.2f}" for v in r] for r in reordered.values],
        texttemplate="%{text}", colorscale=colorscale,
        zmid=0, zmin=-1, zmax=1, colorbar=dict(title="ρ")))
    fig.update_layout(template=THEME, height=680,
                      margin=dict(l=0, r=0, t=50, b=0),
                      title=f"Clustered Correlation (Ward) — 🔴 ρ > {over_thresh}",
                      xaxis_tickangle=-45, shapes=shapes)
    return fig


# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG & CSS
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Quant Dashboard v11", page_icon="📈",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
.main,[data-testid="stAppViewContainer"]{background:#f8f9fb;color:#1a1a2e}
[data-testid="stSidebar"]{background:#ffffff;border-right:1px solid #e0e4ea}
[data-testid="stMetricValue"]{font-size:1.35rem;font-weight:700;color:#1a1a2e}
[data-testid="stMetricLabel"]{font-size:.75rem;color:#5a6070;font-weight:500}
div.stTabs [data-baseweb="tab"]{height:38px;padding:0 18px;border-radius:6px 6px 0 0;
    font-weight:600;color:#3a3f52}
div.stTabs [aria-selected="true"]{background:#fff;color:#0057ff;border-bottom:3px solid #0057ff}
[data-testid="stDataFrame"]{border:1px solid #e0e4ea;border-radius:8px}
[data-testid="stExpander"]{background:#fff;border:1px solid #e0e4ea;border-radius:8px}
hr{border-color:#e0e4ea}
h1,h2,h3{color:#1a1a2e!important}
.sizing-panel{background:#fff;border:1px solid #e0e4ea;border-radius:10px;
    padding:14px 18px 6px;margin-bottom:16px}
.sizing-label{font-size:.78rem;color:#5a6070;font-weight:600;margin-bottom:2px}
.sizing-default{font-size:.70rem;color:#93a1a1}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────

with st.sidebar:
    st.title("⚙️ Settings")
    data_dir_input = st.text_input("Data directory", value=str(DATA_DIR))
    lookback       = st.slider("Lookback (years)", 1, 10, 10)
    st.divider()
    st.caption("**Commission rates ($/contract, round-trip)**")
    COMMISSION_PER_MICRO = st.number_input("Micro ($/contract)",     value=0.88, step=0.05,
        help="Charged per contract, not per trade -- a 3-lot trade costs 3x this. "
             "Default $0.88 is blended from Andrea's real June 2026 statement "
             "(MES/MNQ/MGC fee-lines / contracts traded, volume-weighted). Corrected "
             "2026-07: the previous $1.75 default was ~2x this, calibrated against "
             "the order-history file's flat exchange fee instead of the statement's "
             "real (higher, NFA/broker-inclusive) cost.")
    COMMISSION_PER_MINI  = st.number_input("Mini/Full ($/contract)", value=3.31, step=0.05,
        help="Charged per contract, not per trade. Default $3.31 is blended from "
             "the real June 2026 statement (CL/GC/NQ fee-lines / contracts traded, "
             "volume-weighted). Corrected 2026-07: the previous $6.60 default was "
             "~2x this, same root cause as the micro default above.")
    st.divider()
    st.caption("**Risk Parity**")
    rp_target = st.number_input("Target daily risk/system ($)",
                                 value=500.0, min_value=50.0,
                                 max_value=5000.0, step=50.0)
    rp_max_ct = st.slider("Max contracts (RP)", 1, 30, 20)
    st.divider()
    st.caption("v11.0 — ReadMe names · Dynamic sizing · Raw-trade P&L")


# ─────────────────────────────────────────────────────────────────────────────
# LOAD DATA
# ─────────────────────────────────────────────────────────────────────────────

with st.spinner("Loading trading systems from raw trades…"):
    systems, load_warnings = load_all_systems(
        data_dir_input, COMMISSION_PER_MICRO, COMMISSION_PER_MINI,
        _version=_LOAD_VERSION)

if not systems:
    st.error("No .xlsx files found.")
    st.stop()

for w in load_warnings:
    st.warning(w)

n_sys        = len(systems)
total_trades = sum(len(si["trades"]) for si in systems.values())

st.title(f"📊 Quant Portfolio Dashboard v11  —  {n_sys} systems · {total_trades:,} trades")

_active_stems = [s for s in systems
                 if not systems[s]["trades"].empty and not systems[s]["equity"].empty]

# ── Initialise session_state sizing dict on first load ─────────────────────
if "sizing" not in st.session_state:
    st.session_state["sizing"] = {
        stem: systems[stem]["default_n"] for stem in systems
    }
# Ensure any newly discovered system is present in session state
for stem in systems:
    if stem not in st.session_state["sizing"]:
        st.session_state["sizing"][stem] = systems[stem]["default_n"]


def _cur_ratio(stem: str) -> float:
    """Sizing ratio for this stem: current_n / default_n."""
    default = systems[stem]["default_n"]
    if default == 0:
        return 0.0
    cur = st.session_state["sizing"].get(stem, default)
    return cur / default


# Risk parity (advisory)
_rp_sizing, _rp_detail = compute_risk_parity_sizing(
    systems, _active_stems, lookback,
    target_daily_risk_usd=rp_target, max_contracts=rp_max_ct)

# ── Sidebar Excel export ───────────────────────────────────────────────────
with st.sidebar:
    st.divider()
    _export_corr = None
    _export_eq   = {}
    for stem in _active_stems:
        si = systems[stem]
        eq = get_net_equity_trimmed(si, lookback, _cur_ratio(stem))
        if not eq.empty:
            _export_eq[stem] = eq
    if len(_export_eq) >= 2:
        _corr_df_exp = combine_equity_curves(_export_eq, lookback)
        _export_corr = _corr_df_exp.diff().dropna().corr()

    if st.button("📥 Generate Excel Report", use_container_width=True):
        with st.spinner("Building workbook…"):
            ratios = {s: _cur_ratio(s) for s in systems}
            xls = build_excel_export(systems, lookback,
                                      sizing_ratios=ratios,
                                      corr_matrix=_export_corr)
        st.download_button(
            "⬇️ Download Excel", data=xls,
            file_name=f"quant_v11_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────────────────

def b5_is_none(si: dict) -> bool:
    """True if b5_total is absent → show '—' not '❌' in the overview table."""
    return si["b5_total"] is None


tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "🖥 Systems", "📦 Portfolio",
    "🔬 Correlation & Optimisation", "🔮 Forward Analysis",
    "🔗 Trade Matching"])


# ═════════════════════════════════════════════════════════════════════════════
# TAB 1 — SYSTEMS
# ═════════════════════════════════════════════════════════════════════════════

with tab1:
    # ── SIZING PANEL ─────────────────────────────────────────────────────────
    st.subheader("⚖️ Position Sizing Editor")
    st.caption(
        "Adjust contract counts below. Metrics and equity curves rescale live as "
        "**P&L × (current / default)**. Default values come from ReadMe.txt. "
        "Raw TradeStation P&L is never altered — only the display is scaled.")

    # Reset button
    col_reset, _ = st.columns([1, 5])
    with col_reset:
        if st.button("↺ Reset all to ReadMe defaults", use_container_width=True):
            st.session_state["sizing"] = {
                stem: systems[stem]["default_n"] for stem in systems
            }
            st.rerun()

    # Render a compact grid: 4 systems per row
    stems_list = list(systems.keys())
    n_cols     = 4
    for row_start in range(0, len(stems_list), n_cols):
        row_stems = stems_list[row_start: row_start + n_cols]
        cols      = st.columns(n_cols)
        for col_widget, stem in zip(cols, row_stems):
            si      = systems[stem]
            default = si["default_n"]
            ctype   = si["contract_type"]
            cur_val = st.session_state["sizing"].get(stem, default)

            with col_widget:
                new_val = st.number_input(
                    label=si["display_name"][:30],
                    min_value=0,
                    max_value=50,
                    value=int(cur_val),
                    step=1,
                    key=f"sz_{stem}",
                    help=(f"ReadMe default: {default} {ctype}\n"
                          f"Contract label: {si['contract_label']}"),
                )
                st.caption(f"Default: {default} {ctype}")
                if new_val != cur_val:
                    st.session_state["sizing"][stem] = new_val

    st.divider()

    # ── OVERVIEW TABLE ───────────────────────────────────────────────────────
    st.subheader("System Overview")
    st.caption(
        "**Src:** 📊 = metrics from raw Trades List (correct).  "
        "⚠️ = system absent from ReadMe.txt (filename used).  "
        "All P&L values are TradeStation full-position dollars, scaled by "
        "current sizing ratio.")

    ov_rows = []
    for stem, si in systems.items():
        ratio     = _cur_ratio(stem)
        cur_n     = st.session_state["sizing"].get(stem, si["default_n"])
        src_badge = "📊" if si["data_source"] == "trades" else "⚠️"
        inactive  = " ⏸" if si["is_inactive"] else ""
        m         = compute_metrics(si["trades"], si["comm_per_trade"], ratio)
        ov_rows.append({
            "Src":        src_badge + inactive,
            "System":     si["display_name"],
            "Sym":        si["symbol"],
            "Contracts":  f"{cur_n} {si['contract_type']}  (def {si['default_n']})",
            "# Trades":   m.get("# Trades", len(si["trades"])),
            "Net Profit": f'${m.get("Net Profit ($)", 0):,.0f}' if m else "—",
            "Max DD":     f'${m.get("Max Drawdown ($)", 0):,.0f}' if m else "—",
            "Sharpe":     m.get("Sharpe Ratio", "—"),
            "Calmar":     m.get("Calmar Ratio", "—"),
            "PF":         m.get("Profit Factor", "—"),
            "Win %":      f'{m.get("Win Rate", 0):.1%}' if m else "—",
            "B5✓":        ("✅" if si["b5_match"]
                           else ("—" if b5_is_none(si) else "❌")),
        })

    st.dataframe(pd.DataFrame(ov_rows), use_container_width=True,
                 height=min(42 + 36 * n_sys, 700), hide_index=True)

    with st.expander("🔍 Data quality & validation details"):
        dq_rows = []
        for stem, si in systems.items():
            last_cum = (si["trades"]["cum_pnl"].iloc[-1]
                        if not si["trades"].empty else None)
            dq_rows.append({
                "File":         stem[:45],
                "ReadMe match": "✅" if si["matched"] else "⚠️ no match",
                "Display name": si["display_name"][:40],
                "Contract":     si["contract_label"],
                "Default N":    si["default_n"],
                "Current N":    st.session_state["sizing"].get(stem, si["default_n"]),
                "Comm/contract": f"${si['comm_per_trade']:.2f}",
                "B5 (TS TNP)":  f"${si['b5_total']:,.2f}" if si["b5_total"] else "—",
                "Last cum $":   f"${last_cum:,.2f}" if last_cum is not None else "—",
                "B5 match":     "✅" if si["b5_match"] else "❌",
            })
        st.dataframe(pd.DataFrame(dq_rows), hide_index=True, use_container_width=True)

        st.markdown("**Date ranges per system:**")
        dr_rows = []
        for stem, si in systems.items():
            if not si["trades"].empty:
                dates = si["trades"]["exit_date"].dropna()
                dr_rows.append({
                    "System":   si["display_name"][:40],
                    "First":    dates.min().strftime("%Y-%m-%d"),
                    "Last":     dates.max().strftime("%Y-%m-%d"),
                    "# Trades": len(si["trades"]),
                })
        st.dataframe(pd.DataFrame(dr_rows), hide_index=True, use_container_width=True)

    st.divider()

    # ── SYSTEM EXPLORER ──────────────────────────────────────────────────────
    st.subheader("System Explorer")
    chosen = st.selectbox("Select system", list(systems.keys()),
                          format_func=lambda x: systems[x]["display_name"])

    si_c     = systems[chosen]
    trades_c = si_c["trades"]
    ratio_c  = _cur_ratio(chosen)
    cur_n_c  = st.session_state["sizing"].get(chosen, si_c["default_n"])

    if trades_c.empty:
        st.warning("No trades found for this system.")
    else:
        cutoff   = trades_c["exit_date"].max() - pd.Timedelta(days=int(lookback * 365))
        tr_trim  = trades_c[trades_c["exit_date"] >= cutoff].copy()
        eq_net_c = build_net_equity(tr_trim, si_c["comm_per_trade"], ratio_c)
        if not eq_net_c.empty:
            eq_net_c = eq_net_c - eq_net_c.iloc[0]

        m_c = compute_metrics(tr_trim, si_c["comm_per_trade"], ratio_c)

        st.caption(
            f"**System:** {si_c['display_name']}  |  "
            f"**Contract:** {si_c['contract_label']}  |  "
            f"**Current sizing:** {cur_n_c} × {si_c['contract_type']} "
            f"(default: {si_c['default_n']}, scale: {ratio_c:.2f}×)  |  "
            f"**B5:** {'✅' if si_c['b5_match'] else '❌'}")

        cols5 = st.columns(5)
        for col, (lbl, key, fmt) in zip(cols5, [
            ("Net Profit",   "Net Profit ($)",   "$"),
            ("Max Drawdown", "Max Drawdown ($)",  "$"),
            ("Sharpe",       "Sharpe Ratio",      "f"),
            ("PF",           "Profit Factor",     "f"),
            ("Win Rate",     "Win Rate",          "pct"),
        ]):
            v = m_c.get(key, 0)
            col.metric(lbl,
                       f"${v:,.0f}" if fmt == "$" else
                       (f"{v:.1%}" if fmt == "pct" else f"{v:.2f}"))

        if not eq_net_c.empty:
            color_idx = list(systems.keys()).index(chosen)
            st.plotly_chart(
                plot_equity(eq_net_c, si_c["display_name"],
                            COLORS[color_idx % len(COLORS)]),
                use_container_width=True)
            st.plotly_chart(plot_monthly_heatmap(eq_net_c, si_c["display_name"]),
                            use_container_width=True)

        with st.expander("📋 Trade list"):
            td = tr_trim[["trade_id","entry_date","exit_date",
                           "direction","n_contracts","pnl","cum_pnl"]].copy()
            td.columns = ["#","Entry","Exit","Dir","N Ctrts (TS)",
                          "Raw P&L ($)","Cum P&L ($)"]
            st.dataframe(td, use_container_width=True, height=320)


# ═════════════════════════════════════════════════════════════════════════════
# TAB 2 — PORTFOLIO
# ═════════════════════════════════════════════════════════════════════════════

with tab2:
    st.subheader("Portfolio Builder")
    st.caption(
        "Portfolio equity = sum of per-system net-equity curves scaled by "
        "current sizing ratios (set in Tab 1).")

    selected_systems = st.multiselect(
        "Select systems to include",
        _active_stems,
        default=_active_stems,
        format_func=lambda x: systems[x]["display_name"])

    if not selected_systems:
        st.info("Select at least one system.")
        st.stop()

    # Build portfolio using current sizing from session_state
    # Skip systems where current sizing = 0 (they contribute nothing)
    curves     = {}
    label_map  = {}
    for stem in selected_systems:
        si    = systems[stem]
        ratio = _cur_ratio(stem)
        if ratio == 0.0:
            continue
        eq    = get_net_equity_trimmed(si, lookback, ratio)
        if not eq.empty:
            curves[stem]    = eq
            label_map[stem] = si["display_name"][:30]

    if not curves:
        st.warning("No valid equity curves.")
        st.stop()

    eq_df   = combine_equity_curves(curves, lookback_years=lookback)
    port_eq = eq_df.sum(axis=1)
    pm      = compute_portfolio_metrics(port_eq)

    mc6 = st.columns(6)
    for col, (lbl, key, fmt) in zip(mc6, [
        ("Portfolio P&L",   "Net Profit ($)",   "$"),
        ("Ann. Return",     "Ann. Return ($)",   "$"),
        ("Max Drawdown",    "Max Drawdown ($)",  "$"),
        ("Ann. Volatility", "Ann. Volatility",   "$"),
        ("Sharpe",          "Sharpe Ratio",      "f"),
        ("Calmar",          "Calmar Ratio",      "f"),
    ]):
        v = pm.get(key, 0)
        col.metric(lbl, f"${v:,.0f}" if fmt == "$" else f"{v:.2f}")

    st.plotly_chart(
        plot_portfolio_equity(port_eq, eq_df, label_map=label_map),
        use_container_width=True)

    # Contribution bar
    st.subheader("System Contribution to Portfolio P&L")
    contrib  = {s: eq_df[s].iloc[-1] - eq_df[s].iloc[0] for s in eq_df.columns}
    cs       = pd.Series(contrib).sort_values(ascending=True)
    bar_labels = [systems[k]["display_name"][:38] for k in cs.index]
    fig_bar  = go.Figure(go.Bar(
        x=cs.values, y=bar_labels, orientation="h",
        marker_color=[C["red"] if v < 0 else C["green"] for v in cs.values],
        text=[f"${v:,.0f}" for v in cs.values],
        textposition="outside"))
    fig_bar.update_layout(template=THEME, height=max(300, 38 * len(cs)),
                           margin=dict(l=0, r=70, t=10, b=0),
                           xaxis_title="P&L Contribution ($)")
    st.plotly_chart(fig_bar, use_container_width=True, key="chart_1")

    st.plotly_chart(plot_monthly_heatmap(port_eq, "Portfolio"), use_container_width=True)

    # ── Drawdown Decomposition ────────────────────────────────────────────────
    st.subheader("🔍 Drawdown Decomposition")
    st.caption("For each major portfolio drawdown, which system caused it?")
    dd_episodes = decompose_drawdown(eq_df, port_eq, top_n=3)
    if dd_episodes:
        for ep_i, ep in enumerate(dd_episodes):
            title = (f"DD #{ep_i+1}:  {ep['peak_date'].strftime('%Y-%m-%d')} → "
                     f"{ep['trough_date'].strftime('%Y-%m-%d')}  "
                     f"(Loss: ${ep['dd_abs']:,.0f})")
            with st.expander(title, expanded=(ep_i == 0)):
                blame_s = sorted(ep["blame_pct"].items(),
                                  key=lambda x: x[1], reverse=True)
                blame_s = [(k, v) for k, v in blame_s if v > 0][:8]
                if blame_s:
                    bkeys = [systems[k]["display_name"][:30] if k in systems else k[:30]
                             for k, _ in blame_s]
                    bvals = [v for _, v in blame_s]
                    fig_b = go.Figure(go.Bar(
                        y=bkeys, x=bvals, orientation="h",
                        marker_color=[C["red"] if v > 15 else C["amber"] for v in bvals],
                        text=[f"{v:.1f}%" for v in bvals],
                        textposition="outside"))
                    fig_b.update_layout(template=THEME,
                                         height=max(200, 35 * len(bkeys)),
                                         margin=dict(l=0, r=60, t=10, b=0),
                                         xaxis_title="Blame (%)")
                    st.plotly_chart(fig_b, use_container_width=True, key=f"chart_2_{ep_i}")
    else:
        st.info("No significant drawdown episodes in this lookback window.")

    with st.expander("📉 Overall Drawdown Analysis"):
        dd     = port_eq - port_eq.cummax()
        dd_neg = dd[dd < 0]
        if not dd_neg.empty:
            d1, d2, d3 = st.columns(3)
            d1.metric("Max Drawdown",     f"${dd_neg.min():,.0f}")
            d2.metric("Avg Drawdown",     f"${dd_neg.mean():,.0f}")
            d3.metric("Time in Drawdown", f"{len(dd_neg)/len(dd)*100:.1f}%")
            fig_dd = go.Figure(go.Scatter(
                x=dd.index, y=dd.values, fill="tozeroy",
                line=dict(color=C["dd_line"]), fillcolor=C["drawdown"]))
            fig_dd.update_layout(template=THEME, height=240,
                                  margin=dict(l=0, r=0, t=0, b=0),
                                  yaxis_title="Drawdown ($)")
            st.plotly_chart(fig_dd, use_container_width=True, key="chart_3")


# ═════════════════════════════════════════════════════════════════════════════
# TAB 3 — CORRELATION & OPTIMISATION
# ═════════════════════════════════════════════════════════════════════════════

with tab3:
    all_curves = {}
    for stem in _active_stems:
        si = systems[stem]
        eq = get_net_equity_trimmed(si, lookback, _cur_ratio(stem))
        if not eq.empty:
            all_curves[stem] = eq

    if len(all_curves) < 2:
        st.info("Need ≥ 2 systems with data.")
        st.stop()

    corr_df  = combine_equity_curves(all_curves, lookback_years=lookback)
    daily_r  = corr_df.diff().dropna()
    corr_mat = daily_r.corr()

    # ── Clustered Correlation ─────────────────────────────────────────────────
    st.subheader("Clustered Correlation Heatmap")
    st.caption("Ward-linkage hierarchical clustering. 🔴 boxes = over-exposed pairs.")

    over_exp_thresh = st.slider("Over-exposed threshold (ρ)", 0.5, 0.95, 0.70, 0.05)
    cluster_score   = compute_cluster_risk_score(corr_mat)
    upper           = corr_mat.where(np.triu(np.ones(corr_mat.shape), k=1).astype(bool))
    corr_vals       = upper.stack()
    n_overexp       = (corr_vals > over_exp_thresh).sum()

    c1, c2, c3 = st.columns(3)
    c1.metric("Cluster Risk Score", f"{cluster_score:.0f}/100",
              help="0 = uncorrelated, 100 = all move together")
    c2.metric("Avg Pairwise ρ",     f"{corr_vals.mean():.3f}")
    c3.metric(f"Pairs ρ>{over_exp_thresh}", f"{n_overexp}",
              delta="⚠️ Concentration!" if n_overexp > 3 else "OK",
              delta_color="inverse")

    st.plotly_chart(
        plot_clustered_correlation(corr_mat, systems, over_exp_thresh),
        use_container_width=True)

    high_corr = corr_vals[corr_vals > over_exp_thresh].sort_values(ascending=False)
    if not high_corr.empty:
        with st.expander(f"⚠️ Over-exposed pairs (ρ > {over_exp_thresh})"):
            hc_df = pd.DataFrame({
                "System A":    [systems.get(i[0], {}).get("display_name", i[0])[:38]
                                for i in high_corr.index],
                "System B":    [systems.get(i[1], {}).get("display_name", i[1])[:38]
                                for i in high_corr.index],
                "Correlation": high_corr.values.round(3),
                "Risk":        ["🔴 High" if v > 0.85 else "🟡 Watch"
                                for v in high_corr.values],
            })
            st.dataframe(hc_df, hide_index=True, use_container_width=True)

    st.divider()

    # ── Risk Parity ───────────────────────────────────────────────────────────
    st.subheader("⚖️ Risk Parity Sizing (Advisory)")
    st.caption(
        f"Inverse-volatility weighting targeting **${rp_target:,.0f}/day** per system. "
        "These are advisory — use the Tab 1 editor to apply them.")

    rp_rows = []
    for stem in _active_stems:
        si = systems[stem]
        d  = _rp_detail.get(stem, {})
        rp_rows.append({
            "System":        si["display_name"][:35],
            "Symbol":        si["symbol"],
            "Contract":      si["contract_label"],
            "Daily Vol ($)": f"${d.get('vol_1ct', 0):,.0f}",
            "Raw N":         f'{d.get("raw_n", si["default_n"]):.1f}',
            "RP Advisory N": _rp_sizing.get(stem, si["default_n"]),
            "Current N":     st.session_state["sizing"].get(stem, si["default_n"]),
        })
    st.dataframe(pd.DataFrame(rp_rows), hide_index=True, use_container_width=True)

    st.divider()

    # ── Recommended Sub-Portfolios ────────────────────────────────────────────
    st.subheader("🎯 Recommended Sub-Portfolios")
    st.caption("Subsets selected by different risk/return criteria.")

    reccos = recommend_portfolios(corr_df, systems)
    for rec_i, rec in enumerate(reccos):
        with st.expander(f"{rec['name']} — {rec['description']}"):
            rec_stems = [s for s in rec["systems"] if s in all_curves]
            if not rec_stems:
                st.write("Insufficient data.")
                continue
            ratios_rec = {s: _cur_ratio(s) for s in rec_stems}
            rec_eq     = build_portfolio_equity(systems, rec_stems, lookback, ratios_rec)
            rec_m      = compute_portfolio_metrics(rec_eq)
            mc4 = st.columns(4)
            for col, (lbl, key, fmt) in zip(mc4, [
                ("Net P&L", "Net Profit ($)", "$"), ("Sharpe", "Sharpe Ratio", "f"),
                ("Max DD",  "Max Drawdown ($)", "$"), ("Calmar", "Calmar Ratio", "f"),
            ]):
                v = rec_m.get(key, 0)
                col.metric(lbl, f"${v:,.0f}" if fmt == "$" else f"{v:.2f}")
            if not rec_eq.empty:
                norm = rec_eq - rec_eq.iloc[0]
                fig_r = go.Figure(go.Scatter(x=norm.index, y=norm.values,
                                              line=dict(color=C["blue"], width=2)))
                fig_r.update_layout(template=THEME, height=320,
                                     margin=dict(l=0, r=0, t=10, b=0),
                                     yaxis_title="Cum. P&L ($)")
                st.plotly_chart(fig_r, use_container_width=True, key=f"chart_4_{rec_i}")

    st.divider()

    # ── Risk / Return scatter ─────────────────────────────────────────────────
    st.subheader("Risk / Return Scatter")
    scatter_rows = []
    for stem in corr_df.columns:
        d   = daily_r[stem]
        col = corr_df[stem]
        net   = col.iloc[-1] - col.iloc[0]
        n_yr  = max((col.index[-1] - col.index[0]).days / 365.25, 0.01)
        ann_r = net / n_yr
        ann_v = d.std() * np.sqrt(252)
        sharpe_v = ann_r / ann_v if ann_v > 0 else 0
        max_dd_v = (col - col.cummax()).min()
        scatter_rows.append({
            "System":         systems.get(stem, {}).get("display_name", stem)[:32],
            "Symbol":         systems.get(stem, {}).get("symbol", ""),
            "Sharpe":         round(float(sharpe_v), 2),
            "Max DD ($)":     round(float(max_dd_v), 0),
            "Ann Return ($)": round(float(ann_r), 0),
            "_sz":            max(abs(float(ann_r)), 1.0),
        })
    if scatter_rows:
        sc_df  = pd.DataFrame(scatter_rows)
        fig_sc = px.scatter(sc_df, x="Max DD ($)", y="Sharpe",
                            text="System", color="Symbol",
                            size="_sz", size_max=42, template=THEME,
                            title="Sharpe vs Max Drawdown  (bubble = |Ann. Return|)",
                            hover_data={"Ann Return ($)": True, "_sz": False})
        fig_sc.update_traces(textposition="top center")
        fig_sc.update_layout(height=520, margin=dict(l=0, r=0, t=40, b=0))
        st.plotly_chart(fig_sc, use_container_width=True, key="chart_5")


# ═════════════════════════════════════════════════════════════════════════════
# TAB 4 — FORWARD ANALYSIS
# ═════════════════════════════════════════════════════════════════════════════

with tab4:

    # ── System Health Monitor ─────────────────────────────────────────────────
    st.subheader("🏥 System Health Monitor")
    st.caption("Rolling metrics using ReadMe default sizing (independent of Tab 1 overrides).")

    hw_choice = st.radio("Rolling window", ["3 months (63d)", "6 months (126d)"],
                          horizontal=True, key="hw")
    hw_days   = 63 if "3" in hw_choice else 126

    health_rows   = []
    health_charts = {}
    for stem in _active_stems:
        si  = systems[stem]
        hdf = compute_rolling_health(si["trades"], si["comm_per_trade"], hw_days)
        if hdf.empty:
            health_rows.append({
                "System": si["display_name"][:35], "Status": "⚪",
                "Sharpe": "—", "PF": "—", "Win %": "—", "Period P&L": "—"})
            continue
        health_charts[stem] = hdf
        lt = hdf.iloc[-1]
        tl = health_traffic_light(lt["Sharpe"], lt["Profit Factor"], lt["Win Rate"])
        health_rows.append({
            "System":     si["display_name"][:35],
            "Status":     tl,
            "Sharpe":     f'{lt["Sharpe"]:.2f}',
            "PF":         f'{lt["Profit Factor"]:.2f}',
            "Win %":      f'{lt["Win Rate"]*100:.1f}%',
            "Period P&L": f'${lt["Period P&L"]:,.0f}',
        })

    st.dataframe(pd.DataFrame(health_rows), hide_index=True,
                 use_container_width=True,
                 height=min(42 + 36 * len(health_rows), 520))

    if health_charts:
        with st.expander("📈 Rolling Sharpe over time"):
            fig_rsh = go.Figure()
            for i, (stem, hdf) in enumerate(health_charts.items()):
                fig_rsh.add_trace(go.Scatter(
                    x=hdf.index, y=hdf["Sharpe"].values,
                    name=systems[stem]["display_name"][:25],
                    line=dict(color=COLORS[i % len(COLORS)], width=1.5),
                    hovertemplate="%{x|%Y-%m-%d}<br>Sharpe:%{y:.2f}<extra></extra>"))
            fig_rsh.add_hline(y=0,   line_dash="dot",  line_color=C["zero_line"])
            fig_rsh.add_hline(y=0.8, line_dash="dash", line_color=C["green"],
                              opacity=0.5, annotation_text="Good (0.8)")
            fig_rsh.update_layout(template=THEME, height=420,
                                   margin=dict(l=0, r=0, t=10, b=0),
                                   yaxis_title="Rolling Sharpe",
                                   legend=dict(font=dict(size=10)))
            st.plotly_chart(fig_rsh, use_container_width=True, key="chart_6")

        with st.expander("📈 Rolling Profit Factor over time"):
            fig_rpf = go.Figure()
            for i, (stem, hdf) in enumerate(health_charts.items()):
                if "Profit Factor" in hdf.columns:
                    fig_rpf.add_trace(go.Scatter(
                        x=hdf.index, y=hdf["Profit Factor"].values,
                        name=systems[stem]["display_name"][:25],
                        line=dict(color=COLORS[i % len(COLORS)], width=1.5)))
            fig_rpf.add_hline(y=1.0, line_dash="dot", line_color=C["red"],
                              annotation_text="Break-even (1.0)")
            fig_rpf.update_layout(template=THEME, height=420,
                                   margin=dict(l=0, r=0, t=10, b=0),
                                   yaxis_title="Rolling PF",
                                   legend=dict(font=dict(size=10)))
            st.plotly_chart(fig_rpf, use_container_width=True, key="chart_7")

    st.divider()

    # ── Monte Carlo ───────────────────────────────────────────────────────────
    st.subheader("🎲 Monte Carlo Forward Projection")
    st.caption(
        "Bootstrap resampling of historical daily net returns → confidence bands. "
        "Uses current sizing from Tab 1.")

    mc_c1, mc_c2 = st.columns(2)
    with mc_c1:
        mc_sims = st.selectbox("Simulations", [500, 1000, 2000, 5000], index=1)
    with mc_c2:
        mc_days_opt = st.selectbox(
            "Horizon", [63, 126, 252],
            format_func=lambda x: {63: "3 months", 126: "6 months", 252: "12 months"}[x],
            index=1)

    if st.button("🎲 Run Monte Carlo", type="primary", key="run_mc"):
        with st.spinner(f"Running {mc_sims:,} simulations…"):
            ratios_mc = {s: _cur_ratio(s) for s in _active_stems}
            mc_res = monte_carlo_simulation(
                systems, _active_stems, lookback,
                n_sims=mc_sims, forward_days=mc_days_opt,
                sizing_ratios=ratios_mc)

        if mc_res is None:
            st.warning("Not enough data.")
        else:
            pct     = mc_res["percentiles"]
            days_x  = list(range(len(pct)))
            hor_lbl = {63: "3 months", 126: "6 months",
                       252: "12 months"}.get(mc_days_opt, "")

            fig_mc = go.Figure()
            fig_mc.add_trace(go.Scatter(x=days_x, y=pct["95th"].values,
                                         mode="lines", line=dict(width=0),
                                         showlegend=False))
            fig_mc.add_trace(go.Scatter(x=days_x, y=pct["5th"].values,
                                         mode="lines", line=dict(width=0),
                                         fill="tonexty",
                                         fillcolor="rgba(38,139,210,0.12)",
                                         name="5-95th pct"))
            fig_mc.add_trace(go.Scatter(x=days_x, y=pct["75th"].values,
                                         mode="lines", line=dict(width=0),
                                         showlegend=False))
            fig_mc.add_trace(go.Scatter(x=days_x, y=pct["25th"].values,
                                         mode="lines", line=dict(width=0),
                                         fill="tonexty",
                                         fillcolor="rgba(38,139,210,0.22)",
                                         name="25-75th pct"))
            fig_mc.add_trace(go.Scatter(x=days_x, y=pct["50th"].values,
                                         name="Median",
                                         line=dict(color=C["blue"], width=2.5)))
            fig_mc.add_hline(y=0, line_dash="dot", line_color=C["zero_line"])
            fig_mc.update_layout(template=THEME, height=450,
                                  margin=dict(l=0, r=0, t=30, b=0),
                                  title=f"Monte Carlo: {mc_sims:,} × {hor_lbl}",
                                  xaxis_title="Trading Days Forward",
                                  yaxis_title="Projected P&L ($)")
            st.plotly_chart(fig_mc, use_container_width=True, key="chart_8")

            final = mc_res["paths"].iloc[-1]
            ec    = st.columns(5)
            ec[0].metric("Median",  f"${final.median():,.0f}")
            ec[1].metric("Mean",    f"${final.mean():,.0f}")
            ec[2].metric("5th pct", f"${final.quantile(0.05):,.0f}")
            ec[3].metric("95th",    f"${final.quantile(0.95):,.0f}")
            ec[4].metric("P(loss)", f"{(final < 0).mean():.1%}")

            dd_rows = [{"DD Threshold": f"${t:,.0f}",
                        "Probability":  f"{p:.1%}",
                        "Bar": "█" * int(p * 30)}
                       for t, p in sorted(mc_res["dd_probs"].items())]
            st.dataframe(pd.DataFrame(dd_rows), hide_index=True,
                         use_container_width=True)

    st.divider()

    # ── Sizing Recommendation Summary ─────────────────────────────────────────
    st.subheader("📋 Sizing Recommendation Summary")
    st.caption("Based on rolling health metrics (3-month window). Not financial advice.")

    rec_rows = []
    for stem in _active_stems:
        si  = systems[stem]
        nd  = si["default_n"]
        hdf = compute_rolling_health(si["trades"], si["comm_per_trade"], 63)
        cur_n = st.session_state["sizing"].get(stem, nd)

        if hdf.empty:
            rec_rows.append({
                "System":   si["display_name"][:35],
                "Contract": si["contract_label"],
                "Current N": cur_n,
                "Health":   "⚪ No data",
                "Suggestion": "Insufficient data"})
            continue

        lt  = hdf.iloc[-1]
        sh, pf_v, wr = lt["Sharpe"], lt["Profit Factor"], lt["Win Rate"]
        tl  = health_traffic_light(sh, pf_v, wr)
        if tl == "🟢":
            sugg = ("Scale up — strong momentum" if sh > 1.5 and pf_v > 1.5
                    else "Maintain — performing well")
        elif tl == "🟡":
            sugg = "Monitor — consider reducing if trend continues"
        else:
            sugg = "Consider reducing or pausing" if sh < -0.5 else "Consider reducing"

        rec_rows.append({
            "System":     si["display_name"][:35],
            "Contract":   si["contract_label"],
            "Default N":  nd,
            "Current N":  cur_n,
            "Health":     f"{tl} Sh={sh:.1f} PF={pf_v:.1f}",
            "Suggestion": sugg,
        })

    st.dataframe(pd.DataFrame(rec_rows), hide_index=True,
                 use_container_width=True,
                 height=min(42 + 36 * len(rec_rows), 520))
    st.info("⚠️ Suggestions based on rolling historical metrics. "
            "Not financial advice. Past performance ≠ future results.")

    st.divider()

    # ── Regime Analysis ───────────────────────────────────────────────────────
    st.subheader("🌡️ Regime Analysis")
    st.caption("Classifies market conditions via portfolio realized volatility.")

    reg_c1, reg_c2, reg_c3 = st.columns(3)
    with reg_c1:
        reg_vw = st.selectbox("Vol window (days)", [10, 15, 20, 30, 40], index=2, key="r_vw")
    with reg_c2:
        reg_lp = st.slider("Low-vol pct", 10, 45, 33, key="r_lp")
    with reg_c3:
        reg_hp = st.slider("High-vol pct", 55, 90, 66, key="r_hp")

    _reg_curves = {}
    for stem in _active_stems:
        si = systems[stem]
        eq = get_net_equity_trimmed(si, lookback, _cur_ratio(stem))
        if not eq.empty:
            _reg_curves[stem] = eq

    if len(_reg_curves) < 2:
        st.info("Need ≥ 2 active systems.")
    else:
        regime_df, regime_thresholds = compute_regime_series(
            _reg_curves, vol_window=reg_vw, low_pct=reg_lp, high_pct=reg_hp)

        if regime_df.empty:
            st.warning("Not enough data for regime classification.")
        else:
            cur_regime = regime_df["regime"].iloc[-1]
            reg_emoji  = {"Low Vol": "😴", "Medium Vol": "😐", "High Vol": "🔥"}
            st.markdown(
                f"### Current Regime: {reg_emoji.get(cur_regime, '')} **{cur_regime}**")

            rm1, rm2, rm3, rm4 = st.columns(4)
            rm1.metric("Current Ann. Vol", f"${regime_df['vol'].iloc[-1]:,.0f}")
            rm2.metric("Thresholds",
                       f"${regime_thresholds['low']:,.0f} / ${regime_thresholds['high']:,.0f}")
            rc  = regime_df["regime"].value_counts()
            tot = len(regime_df)
            rm3.metric("% Low Vol",  f"{rc.get('Low Vol',  0)/tot*100:.0f}%")
            rm4.metric("% High Vol", f"{rc.get('High Vol', 0)/tot*100:.0f}%")

            rcmap = {"Low Vol": C["green"], "Medium Vol": C["amber"], "High Vol": C["red"]}
            fig_reg = go.Figure()
            for label, color in rcmap.items():
                mask = regime_df["regime"] == label
                sub  = regime_df[mask]
                if not sub.empty:
                    fig_reg.add_trace(go.Scatter(
                        x=sub.index, y=sub["vol"].values, mode="markers",
                        marker=dict(color=color, size=3, opacity=0.6), name=label))
            fig_reg.add_hline(y=regime_thresholds["low"],  line_dash="dash",
                              line_color=C["green"], opacity=0.5)
            fig_reg.add_hline(y=regime_thresholds["high"], line_dash="dash",
                              line_color=C["red"], opacity=0.5)
            fig_reg.update_layout(template=THEME, height=300,
                                   margin=dict(l=0, r=0, t=10, b=0),
                                   yaxis_title="Ann. Vol ($)",
                                   legend=dict(orientation="h", y=1.05))
            st.plotly_chart(fig_reg, use_container_width=True, key="chart_9")

            with st.expander("🔄 Regime Transition Probabilities"):
                trans_df = compute_regime_transitions(regime_df)
                if not trans_df.empty:
                    st.caption("P(next = column | today = row)")
                    fig_t = go.Figure(go.Heatmap(
                        z=trans_df.values,
                        x=trans_df.columns.tolist(), y=trans_df.index.tolist(),
                        text=[[f"{v:.1%}" for v in row] for row in trans_df.values],
                        texttemplate="%{text}",
                        colorscale="Blues", zmin=0, zmax=1,
                        colorbar=dict(title="P")))
                    fig_t.update_layout(template=THEME, height=280,
                                         margin=dict(l=0, r=0, t=10, b=0))
                    st.plotly_chart(fig_t, use_container_width=True, key="chart_10")
                    persist = {r: trans_df.loc[r, r] for r in trans_df.index}
                    st.markdown(
                        f"**Persistence:** Low Vol stays {persist.get('Low Vol',0):.0%} · "
                        f"Medium stays {persist.get('Medium Vol',0):.0%} · "
                        f"High Vol stays {persist.get('High Vol',0):.0%}")

            st.markdown("##### System Performance by Regime")
            rp_table = []
            for stem in _active_stems:
                si   = systems[stem]
                cutoff = (si["equity"].index.max() - pd.Timedelta(days=int(lookback * 365))
                          if not si["equity"].empty else pd.Timestamp("2000-01-01"))
                tr_t = (si["trades"][si["trades"]["exit_date"] >= cutoff]
                        if not si["trades"].empty else si["trades"])
                rp = compute_system_regime_performance(
                    tr_t, regime_df, si["comm_per_trade"])
                for rl in ["Low Vol", "Medium Vol", "High Vol"]:
                    m = rp.get(rl, {})
                    if m.get("n_trades", 0) == 0:
                        continue
                    rp_table.append({
                        "System":    si["display_name"][:32],
                        "Regime":    rl,
                        "# Trades":  m["n_trades"],
                        "Total P&L": f'${m["total_pnl"]:,.0f}',
                        "Avg P&L":   f'${m["avg_pnl"]:,.0f}',
                        "Win Rate":  f'{m["win_rate"]:.0%}',
                        "PF":        f'{m["profit_factor"]:.2f}',
                    })
            if rp_table:
                st.dataframe(pd.DataFrame(rp_table), hide_index=True,
                             use_container_width=True,
                             height=min(42 + 36 * len(rp_table), 520))
            st.info("💡 Systems robust across all regimes are the strongest core holdings.")


# ═════════════════════════════════════════════════════════════════════════
# TAB 5 — TRADE MATCHING  (theoretical XLSX trades vs. actual TradeStation
# executions). Uses tradestation_matcher.py — see that module's docstring
# for the data-quality caveats this tab surfaces (order-history coverage
# gaps, no per-strategy tag on fills, continuous-contract price offsets).
# ═════════════════════════════════════════════════════════════════════════

with tab5:
    st.subheader("🔗 Trade Matching — Theoretical (XLSX) vs Actual (TradeStation)")
    st.caption(
        "Matches each system's theoretical trades against real executions from a "
        "TradeStation order-history export, within date/price tolerance. The monthly "
        "PDF statement is used only as an independent P&L/commission cross-check "
        "(it has no per-trade prices, so it can't be used for matching itself). "
        "Strategy trade files can be .xlsx or .csv (flat report export) — drop either "
        "into the Data directory set in the sidebar; if both exist for the same system, "
        "the .csv is used.")

    if tsm is None:
        st.error(
            "`tradestation_matcher.py` could not be imported "
            f"({_TSM_IMPORT_ERROR}). Make sure the file is in the same folder as "
            "app.py, and that `openpyxl`/`pdfplumber` are installed.")
    else:
        OVERRIDE_LOG_PATH = str(DATA_DIR / "trade_overrides_log.csv")

        with st.expander("⚙️ Matching configuration", expanded=True):
            c1, c2, c3 = st.columns(3)
            with c1:
                order_hist_path = st.text_input(
                    "Order-history file OR folder path (.xlsx or .ods)",
                    value=str(DATA_DIR / "trade_data" / "order_history"),
                    key="tsm_order_hist_path",
                    help="TradeStation order/execution export (Entered, Filled/Canceled, "
                         "Symbol, Type, Qty Filled, Filled Price, Order Status, Commission, "
                         "Contract Exp Date). Both .xlsx and .ods exports are supported. "
                         "Point this at a FOLDER (e.g. trade_data/order_history/) to always "
                         "auto-use the most recently added file in it — no need to retype a "
                         "filename each time you drop in a fresh export. Or point it at one "
                         "specific file if you want to pin an exact export.")
            with c2:
                stmt_pdf_path = st.text_input(
                    "Monthly statement PDF path (optional)",
                    value=str(DATA_DIR / "trade_data" / "statements"),
                    key="tsm_pdf_path",
                    help="Used only for the reconciliation check, not trade matching. "
                         "Point this at a FOLDER (e.g. trade_data/statements/) and it will "
                         "pick the PDF whose filename contains the selected year+month "
                         "(TradeStation's own naming, e.g. '..._202601.pdf' for Jan 2026). "
                         "If no filename matches, it falls back to the newest PDF in the "
                         "folder and flags that in the reconciliation section. Or point it "
                         "at one specific file if you want to pin an exact statement.")
            with c3:
                cM, cY = st.columns(2)
                with cM:
                    match_month = st.selectbox("Month", list(range(1, 13)),
                                                index=datetime.now().month - 1,
                                                format_func=lambda m: datetime(2000, m, 1).strftime("%b"),
                                                key="tsm_month")
                with cY:
                    match_year = st.number_input("Year", value=datetime.now().year,
                                                  min_value=2015, max_value=2100, step=1,
                                                  key="tsm_year")

            c4, c5 = st.columns(2)
            with c4:
                date_tol = st.slider("Time tolerance (± minutes)", 0, 60, 10, key="tsm_date_tol",
                                      help="Theoretical entry/exit timestamps are candle-CLOSE "
                                           "times, so the real execution can lag by up to one "
                                           "candle (~2.5-5 min on a 5-min ORB signal, confirmed "
                                           "on real June 2026 fills). Theoretical times are also "
                                           "shifted +7h automatically to align with Chicago fill "
                                           "time — this slider only covers the remaining small "
                                           "execution lag, not timezone. Was a 1-business-day "
                                           "window before; that was masking the (now-fixed) "
                                           "timezone gap and, once fixed, became too wide — it let "
                                           "unrelated same-day fills collide as false matches.")
            with c5:
                tick_tol = st.slider("Price tolerance (± ticks)", 0, 10, 2, key="tsm_tick_tol")

            match_portfolio_tab = st.checkbox(
                "Apply current Tab 1 sizing + commission (match Portfolio tab numbers)",
                value=True, key="tsm_match_portfolio",
                help="When on, 'Theoretical P&L' is scaled by each system's current "
                     "Tab 1 sizing ratio and reduced by the sidebar's commission rate — "
                     "the exact same formula the Portfolio tab uses — so the totals "
                     "below are directly comparable to what Tab 2 shows for this month. "
                     "When off, 'Theoretical P&L' is the raw, unscaled number straight "
                     "from the trade files.")

            run_clicked = st.button("▶️ Run matching engine", type="primary", key="tsm_run_btn")

        if run_clicked:
            if not os.path.exists(order_hist_path):
                st.error(f"Order-history file not found: {order_hist_path}")
            else:
                pdf_arg = stmt_pdf_path if os.path.exists(stmt_pdf_path) else None
                if pdf_arg is None:
                    st.info("No statement PDF/folder found at that path — running without "
                             "the reconciliation cross-check.")
                sizing_ratios_arg   = None
                commission_rates_arg = None
                if match_portfolio_tab:
                    sizing_ratios_arg    = {s: _cur_ratio(s) for s in systems}
                    commission_rates_arg = {s: systems[s]["comm_per_trade"] for s in systems}
                with st.spinner("Matching theoretical trades to actual executions…"):
                    cfg = tsm.MatchConfig(date_tol_minutes=date_tol, tick_tol=tick_tol)
                    tsm_result = tsm.run_matching_engine(
                        str(DATA_DIR), _README_PATH, order_hist_path,
                        int(match_year), int(match_month), pdf_arg, cfg,
                        sizing_ratios_arg, commission_rates_arg)
                st.session_state["tsm_result"] = tsm_result

        tsm_result = st.session_state.get("tsm_result")

        if tsm_result is None:
            st.info("Set the file paths above and click **Run matching engine** to begin.")
        else:
            cov = tsm_result["coverage"]
            month_label = datetime(int(match_year), int(match_month), 1).strftime("%B %Y")

            # ── Data coverage banner — surfaces gaps rather than hiding them ──
            if cov["n"] == 0:
                st.error("Order-history file has no recognizable 'Filled' rows.")
            else:
                cov_lo, cov_hi = cov["min"], cov["max"]
                month_start = pd.Timestamp(int(match_year), int(match_month), 1)
                month_end = month_start + pd.offsets.MonthEnd(1)
                if cov_lo > month_start or cov_hi < month_end:
                    st.warning(
                        f"⚠️ Order-history file covers **{cov_lo:%Y-%m-%d} → {cov_hi:%Y-%m-%d}** "
                        f"({cov['n']:,} filled orders), which does not fully cover "
                        f"**{month_label}** ({month_start:%Y-%m-%d} → {month_end:%Y-%m-%d}). "
                        "Trades outside the covered window will show as unmatched/unreconciled "
                        "— that's a data gap, not a real discrepancy.")
                else:
                    st.success(
                        f"✅ Order-history file fully covers {month_label} "
                        f"({cov_lo:%Y-%m-%d} → {cov_hi:%Y-%m-%d}, {cov['n']:,} filled orders).")

            # ── Systems excluded entirely (not in ReadMe, or ReadMe size is 0) ──
            # Per Andrea: a system with no ReadMe entry or an explicit 0-contract
            # allocation isn't part of the real portfolio, so its trades should
            # count as zero -- NOT show up as large phantom "unmatched" P&L for a
            # system she isn't actually running. load_all_theoretical_trades()
            # drops these before matching even starts; shown here so an excluded
            # system is still visible (in case it's actually a rename/typo that
            # needs fixing in ReadMe.txt, not a system that should stay excluded).
            excluded = tsm_result.get("excluded_systems", [])
            if excluded:
                with st.expander(f"🚫 {len(excluded)} system file(s) excluded from all totals below"):
                    st.caption(
                        "Not in ReadMe.txt, or listed there with 0 contracts — treated as not "
                        "part of the real portfolio, so none of their trades count toward "
                        "Theoretical P&L / Unmatched Impact / Portfolio Totals. If one of these "
                        "is actually an active system (e.g. a rename ReadMe.txt hasn't caught "
                        "up with), add/fix its ReadMe.txt line and it'll be picked up normally.")
                    st.dataframe(pd.DataFrame(excluded), hide_index=True, use_container_width=True)

            # ── Portfolio totals — the number that actually answers "where did
            # the money go", but ONLY once (most of) the 18 systems are loaded ──
            n_loaded = tsm_result.get("n_systems_loaded", 0)
            totals = tsm_result.get("portfolio_totals", {})
            st.subheader(f"💰 Portfolio Totals — {month_label}")
            if n_loaded < 18:
                st.warning(
                    f"⚠️ Only **{n_loaded} of 18 systems** have a trade file loaded right now "
                    f"(found in the Data directory). These totals only cover those "
                    f"{n_loaded} — they will UNDERSTATE the full portfolio gap until all "
                    f"18 systems have a .csv or .xlsx in place. Don't treat this as proof "
                    f"of anything until it's complete.")
            if not totals:
                st.caption("No theoretical trades exited in this month for any loaded system.")
            else:
                # ── Statement P&L/fee, pulled up here as the headline "how much
                # money do I actually have" figure — per Andrea: "consider the
                # monthly statements as the bible... the rest is less important."
                # The matching-engine's own "Actual P&L (net)" below only sums
                # MATCHED + pending-override trades (107 of ~314 June fills);
                # it structurally excludes unmatched actual fills (real P&L
                # that never got attributed to a theoretical signal) and any
                # system whose trade file isn't loaded yet, so it will never
                # equal the statement and shouldn't be read as "cash on hand."
                _recon = tsm_result.get("reconciliation")
                _stmt_row = None
                if _recon is not None and not _recon.empty and "TOTAL" in _recon["root"].values:
                    _stmt_row = _recon[_recon["root"] == "TOTAL"].iloc[0]
                if _stmt_row is not None:
                    st.metric("💵 Statement P&L (real, all fills, ground truth)",
                              f"${_stmt_row['statement_pnl']:,.2f}",
                              help="From the broker statement PDF — the actual cash result "
                                   "for the month, covering every fill on these 6 roots "
                                   "(matched, unmatched, and any system not yet loaded here). "
                                   "This is the number to trust for 'how much do I have.'")
                    st.caption(
                        f"Statement fee: ${_stmt_row['statement_fee']:,.2f} — see the fee "
                        f"explanation in the Reconciliation section below for why this is "
                        f"higher than the order-history-reconstructed commission figure.")
                else:
                    st.caption("No statement PDF reconciled for this month — showing the "
                               "matching-engine's own totals below only (matched + pending-"
                               "override trades, NOT the full real P&L for the month).")
                st.caption("Matching-engine totals below (matched + pending-override trades "
                           "only — a subset of the statement total above):")
                tcol1, tcol2, tcol3 = st.columns(3)
                tcol1.metric("Total Theoretical P&L", f"${totals.get('Theoretical P&L', 0):,.0f}",
                             help="Matches the Portfolio tab's methodology if the sizing/"
                                  "commission toggle above was on when you ran this.")
                tcol2.metric("Total Actual P&L (net, matched trades only)",
                             f"${totals.get('Actual P&L', 0):,.0f}",
                             help="Only matched + pending-override trades — NOT the full "
                                  "month's real P&L. See the Statement P&L figure above for that.")
                tcol3.metric("Total Delta", f"${totals.get('Delta', 0):,.0f}",
                             delta=f"${totals.get('Delta', 0):,.0f}")
                dcol1, dcol2, dcol3, dcol4 = st.columns(4)
                dcol1.metric("Commissions", f"${totals.get('Commissions ($)', 0):,.0f}")
                dcol2.metric("Slippage ($)", f"${totals.get('Slippage ($)', 0):,.0f}",
                             help="Dollar P&L gap (price gap x point value x contracts), "
                                  "not a points/ticks figure -- a small, tolerable price gap "
                                  "on a high-point-value root like GC or NQ can still show as "
                                  "a large dollar number here.")
                dcol3.metric("Override Impact ($)", f"${totals.get('Override Impact ($)', 0):,.0f}")
                dcol4.metric("Unmatched Impact ($)", f"${totals.get('Unmatched Impact ($)', 0):,.0f}")
                st.caption(
                    f"{totals.get('# Matched', 0)} matched · "
                    f"{totals.get('# Overrides (pending review)', 0)} pending override review · "
                    f"{totals.get('# Unmatched Theo', 0)} unmatched theoretical trades, "
                    f"across {n_loaded} system(s).")

            # ── Price-offset calibration transparency ──
            # price_offsets is dict[stem] -> LIST of segments (one per
            # contract-month the system traded, e.g. pre-roll/post-roll --
            # see calibrate_price_offsets()), not a single dict per system,
            # since a mid-month contract roll needs its own offset on each
            # side of the roll.
            offsets = tsm_result.get("price_offsets", {})
            off_rows = []
            for stem, segments in offsets.items():
                for seg in segments:
                    off_rows.append({
                        "System": stem,
                        "Contract": (f"{seg['contract_key'][0]}"
                                     f"{seg['contract_key'][1]:02d}/{seg['contract_key'][2]}"
                                     if seg.get("contract_key") else ""),
                        "Offset (pts)": round(seg["offset"] if seg["applied"] else seg.get("raw_offset", 0.0), 2),
                        "Trusted (auto-matched)": seg["applied"],
                        "# Anchors": seg["n_anchors"],
                        "Spread (pts)": round(seg["spread"], 3),
                        "Date range": f"{seg['date_min']:%Y-%m-%d} → {seg['date_max']:%Y-%m-%d}",
                    })
            if off_rows:
                n_trusted = sum(1 for r in off_rows if r["Trusted (auto-matched)"])
                with st.expander(f"📐 Price offset calibrated for {len(off_rows)} system/contract "
                                  f"segment(s) ({n_trusted} trusted for auto-matching)", expanded=False):
                    st.caption(
                        "Theoretical prices come from a back-adjusted continuous contract, offset "
                        "from the real traded contract — recalibrated separately per contract-month "
                        "a system actually traded, since a mid-month roll (e.g. MNQ M26→U26) resets "
                        "the offset. 'Trusted' segments were tight enough to auto-match on; "
                        "untrusted ones are still offered as override-review candidates (never "
                        "silently auto-matched) rather than discarded.")
                    st.dataframe(pd.DataFrame(off_rows), hide_index=True, use_container_width=True)

            st.divider()

            # ── Per-system breakdown ──
            st.subheader(f"📊 System Breakdown — {month_label}")
            breakdown = tsm_result["breakdown"]
            if breakdown.empty:
                st.info("No theoretical trades exited in this month for any loaded system.")
            else:
                st.dataframe(breakdown, hide_index=True, use_container_width=True,
                             height=min(42 + 36 * len(breakdown), 500))

                fig_bd = go.Figure()
                fig_bd.add_trace(go.Bar(x=breakdown["System"], y=breakdown["Theoretical P&L"],
                                         name="Theoretical P&L", marker_color=C["blue"]))
                fig_bd.add_trace(go.Bar(x=breakdown["System"], y=breakdown["Actual P&L"],
                                         name="Actual P&L", marker_color=C["portfolio"]))
                fig_bd.update_layout(template=THEME, barmode="group", height=380,
                                      margin=dict(l=0, r=0, t=30, b=0),
                                      title="Theoretical vs Actual P&L by system",
                                      yaxis_title="$")
                st.plotly_chart(fig_bd, use_container_width=True, key="chart_tsm_breakdown_bar")

                fig_delta = go.Figure()
                comp_cols = ["Commissions ($)", "Slippage ($)", "Override Impact ($)", "Unmatched Impact ($)"]
                comp_colors = [C["amber"], C["red"], C["green"], C["peak"]]
                for i, (col, color) in enumerate(zip(comp_cols, comp_colors)):
                    fig_delta.add_trace(go.Bar(x=breakdown["System"], y=breakdown[col],
                                                name=col, marker_color=color))
                fig_delta.update_layout(template=THEME, barmode="relative", height=380,
                                         margin=dict(l=0, r=0, t=30, b=0),
                                         title="Delta decomposition by system (sums to Actual − Theoretical)",
                                         yaxis_title="$")
                st.plotly_chart(fig_delta, use_container_width=True, key="chart_tsm_delta_bar")

            st.divider()

            # ── Matched trades ──
            matched = tsm_result["matched"]
            with st.expander(f"✅ Matched trades ({len(matched)})", expanded=False):
                if matched.empty:
                    st.caption("No confident matches within tolerance.")
                else:
                    flagged = matched[matched["had_alt_candidates"]]
                    if not flagged.empty:
                        st.warning(
                            f"{len(flagged)} of these had more than one similarly-good candidate "
                            "fill — double-check them (column `had_alt_candidates`).")
                    if "high_slippage" in matched.columns:
                        slippy = matched[matched["high_slippage"]]
                        if not slippy.empty:
                            st.info(
                                f"{len(slippy)} matched at essentially the same time as the "
                                "signal but with a bigger price gap than the base tolerance — "
                                "treated as slippage, not a manual override (column "
                                "`high_slippage`). NQ/MNQ allow up to 13pts, ES/MES up to 4pts "
                                "before this flag no longer applies at all and it falls to "
                                "override review instead.")
                    st.dataframe(matched.drop(columns=["theo_id"]), hide_index=True,
                                 use_container_width=True, height=min(42 + 36 * len(matched), 600))

            # ── Override candidates — interactive human review ──
            overrides = tsm_result["override_candidates"]
            st.subheader(f"✏️ Manual override review ({len(overrides)} pending)")
            if overrides.empty:
                st.caption("No trades matched outside tight tolerance this month.")
            else:
                st.caption(
                    "These matched on symbol/qty/side within a wider window, but price or date "
                    "fell outside the normal tolerance — consistent with an early/late manual "
                    "exit. No override history exists yet, so nothing is assumed: confirm or "
                    "correct the reason for each, and it's saved to the override log below.")
                override_log = tsm.load_override_log(OVERRIDE_LOG_PATH)
                for i, row in overrides.reset_index(drop=True).iterrows():
                    theo_id = row["theo_id"]
                    prior = override_log[override_log["theo_id"] == theo_id] if not override_log.empty else pd.DataFrame()
                    prior_reason = prior["override_reason"].iloc[0] if not prior.empty else ""
                    with st.container():
                        st.markdown(
                            f"**{row['system_name']}** — {row['symbol']} × {row['n_contracts']} "
                            f"({row['direction']}) — theo P&L ${row['theo_pnl']:,.0f} vs "
                            f"actual net ${row['actual_pnl_net']:,.0f}  \n"
                            f"Entry: theo {row['theo_entry_price']:.2f} @ {row['theo_entry_date']:%Y-%m-%d %H:%M} "
                            f"→ actual {row['actual_entry_price']:.2f} @ {row['actual_entry_time']:%Y-%m-%d %H:%M} "
                            f"({row['entry_diff_ticks']:+.1f} ticks)  \n"
                            f"Exit: theo {row['theo_exit_price']:.2f} @ {row['theo_exit_date']:%Y-%m-%d %H:%M} "
                            f"→ actual {row['actual_exit_price']:.2f} @ {row['actual_exit_time']:%Y-%m-%d %H:%M} "
                            f"({row['exit_diff_ticks']:+.1f} ticks)")
                        rcol1, rcol2 = st.columns([3, 1])
                        with rcol1:
                            reason = st.text_input(
                                "Override reason", value=prior_reason,
                                key=f"tsm_override_reason_{i}_{theo_id}_{row['system_stem']}",
                                placeholder="e.g. exited early on discretionary risk call")
                        with rcol2:
                            if st.button("💾 Save", key=f"tsm_override_save_{i}_{theo_id}_{row['system_stem']}",
                                         use_container_width=True):
                                entry = row.to_dict()
                                entry["override_reason"] = reason
                                entry["match_type"] = "override"
                                entry["reviewed_by"] = "Andrea"
                                entry["reviewed_at"] = datetime.now().isoformat(timespec="seconds")
                                tsm.save_override_entry(OVERRIDE_LOG_PATH, entry)
                                st.success("Saved.")
                        st.divider()

            # ── Unmatched / out-of-scope ──
            uc1, uc2, uc3 = st.columns(3)
            with uc1:
                unmatched_theo = tsm_result["unmatched_theoretical"]
                with st.expander(f"❔ Unmatched theoretical ({len(unmatched_theo)})"):
                    st.caption("Strategy signal with no corresponding execution found — "
                               "either not taken, or outside the order-history's covered dates. "
                               "Its theoretical P&L still counts toward this system's totals "
                               "above (that's what drives the 'Unmatched Impact' column) — "
                               "tagging a reason below is just for your own record, it doesn't "
                               "change the numbers.")
                    if not unmatched_theo.empty:
                        by_sys = (unmatched_theo.groupby("system_name")
                                  .agg(n_trades=("pnl", "size"), total_theo_pnl=("pnl", "sum"))
                                  .sort_values("n_trades", ascending=False).reset_index())
                        by_sys["total_theo_pnl"] = by_sys["total_theo_pnl"].round(2)
                        st.caption("By system (which ones are driving this count):")
                        st.dataframe(by_sys, hide_index=True, use_container_width=True)
                        st.caption("Full list:")
                        st.dataframe(
                            unmatched_theo[["system_name", "entry_date", "entry_price",
                                            "exit_date", "exit_price", "n_contracts", "pnl"]],
                            hide_index=True, use_container_width=True)
                        skip_log = tsm.load_override_log(OVERRIDE_LOG_PATH)
                        st.caption("Tag why a specific signal wasn't taken (optional):")
                        for j, urow in unmatched_theo.reset_index(drop=True).iterrows():
                            u_theo_id = urow["theo_id"]
                            u_prior = skip_log[skip_log["theo_id"] == u_theo_id] if not skip_log.empty else pd.DataFrame()
                            u_prior_reason = (u_prior["override_reason"].iloc[0]
                                               if not u_prior.empty and "override_reason" in u_prior.columns
                                               else "")
                            ucol1, ucol2 = st.columns([3, 1])
                            with ucol1:
                                skip_reason = st.text_input(
                                    f"{urow['system_name']} — {urow['entry_date']:%Y-%m-%d} "
                                    f"(theo P&L ${urow['pnl']:,.0f})",
                                    value=u_prior_reason,
                                    key=f"tsm_skip_reason_{j}_{u_theo_id}_{urow['system_stem']}",
                                    placeholder="e.g. system was off, deliberately skipped")
                            with ucol2:
                                st.write("")
                                if st.button("💾 Save", key=f"tsm_skip_save_{j}_{u_theo_id}_{urow['system_stem']}",
                                             use_container_width=True):
                                    u_entry = urow.to_dict()
                                    u_entry["theo_pnl"] = urow.get("pnl")
                                    u_entry["actual_entry_time"] = ""
                                    u_entry["actual_entry_price"] = ""
                                    u_entry["actual_exit_time"] = ""
                                    u_entry["actual_exit_price"] = ""
                                    u_entry["actual_pnl_net"] = 0.0
                                    u_entry["entry_diff_ticks"] = ""
                                    u_entry["exit_diff_ticks"] = ""
                                    u_entry["override_reason"] = skip_reason
                                    u_entry["match_type"] = "skipped"
                                    u_entry["reviewed_by"] = "Andrea"
                                    u_entry["reviewed_at"] = datetime.now().isoformat(timespec="seconds")
                                    tsm.save_override_entry(OVERRIDE_LOG_PATH, u_entry)
                                    st.success("Saved.")
            with uc2:
                unmatched_actual = tsm_result["unmatched_actual"]
                with st.expander(f"❔ Unmatched actual fills ({len(unmatched_actual)})"):
                    st.caption("Real executions (in the 18 systems' symbols) that no "
                               "theoretical trade claimed — extra/manual trades, or a "
                               "system not currently loaded.")
                    if not unmatched_actual.empty:
                        st.dataframe(
                            unmatched_actual[["root", "side", "qty_filled", "price",
                                              "fill_time", "commission"]].sort_values("fill_time"),
                            hide_index=True, use_container_width=True, height=300)
            with uc3:
                out_scope = tsm_result["out_of_scope_actual"]
                with st.expander(f"🚫 Out-of-scope fills ({len(out_scope)})"):
                    st.caption("Symbols outside the 18 systems (e.g. RTY, HG/MHG) — shown "
                               "for visibility only, never matched.")
                    if not out_scope.empty:
                        st.dataframe(
                            out_scope[["root", "side", "qty_filled", "price", "fill_time"]]
                            .sort_values("fill_time"),
                            hide_index=True, use_container_width=True, height=300)

            st.divider()

            # ── Reconciliation vs monthly statement ──
            st.subheader("🧾 Reconciliation vs monthly statement")
            recon = tsm_result["reconciliation"]
            resolved_pdf = tsm_result.get("resolved_pdf_path")
            if recon.empty:
                if resolved_pdf:
                    st.caption(f"Found `{resolved_pdf}` but couldn't parse/reconcile it — "
                               "reconciliation skipped.")
                else:
                    st.caption("No statement PDF provided (or none found at the given path) — "
                               "reconciliation skipped.")
            else:
                if not tsm_result.get("pdf_is_stamped_match", True):
                    st.warning(
                        f"No PDF filename in the statements folder matched "
                        f"{int(match_year)}-{int(match_month):02d} — falling back to the "
                        f"most recently added PDF instead: `{resolved_pdf}`. These numbers "
                        "may be for the wrong month; add the correct month's statement PDF "
                        "to be sure.")
                else:
                    st.caption(f"Statement used: `{resolved_pdf}`")
                st.caption(
                    "Coarse per-symbol total check: the PDF statement's daily P&L/commission "
                    "totals vs. the order-history reconstruction, for the whole month. This "
                    "does not use trade-level matching (the PDF doesn't support it — see the "
                    "module docstring) so treat P&L mismatches as a prompt to double-check the "
                    "order-history file's coverage, not as a matching error.")
                st.caption(
                    "⚠️ `fee_match` will almost always read False and `fills_fee_reconstructed` "
                    "will almost always run below `statement_fee` — confirmed not fixable from "
                    "this data: the order-history file's Commission column only carries the "
                    "exchange/clearing fee attached at execution (a flat $0.40/contract on "
                    "micros, $1.00/contract on full-size, checked across an entire month with "
                    "zero variance), while the statement's fee total also includes NFA/broker "
                    "charges that never appear in that export. Trust `statement_fee` as the "
                    "real cost — it's what the sidebar's commission defaults are calibrated "
                    "from.")
                st.dataframe(recon, hide_index=True, use_container_width=True)

            st.divider()
            with st.expander("📜 Override log (all systems, all months)"):
                full_log = tsm.load_override_log(OVERRIDE_LOG_PATH)
                if full_log.empty:
                    st.caption("No overrides saved yet.")
                else:
                    st.dataframe(full_log, hide_index=True, use_container_width=True)
