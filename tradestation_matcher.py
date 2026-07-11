"""
tradestation_matcher.py
========================
Trade-matching engine: theoretical strategy trades (from the 18 per-system
XLSX backtest/report files) vs. actual executions (from a TradeStation
order-history export).

─────────────────────────────────────────────────────────────────────────────
WHY THIS MODULE LOOKS THE WAY IT DOES  (read before changing the matching
logic — these are lessons learned from the real sample data, not guesses)
─────────────────────────────────────────────────────────────────────────────

1. The monthly PDF "Commodity Statement" (e.g. 210PYV33_202602.pdf) does NOT
   contain per-trade entry/exit prices. It only has one aggregated P&L line
   per contract-description per day (e.g. "15 15 MAR 26 CME MICRO S&P ...
   P&L US 1,027.50"), which can bundle together closes from several
   different systems that happen to close the same symbol the same day.
   => The PDF is used ONLY for a total P&L / commission reconciliation
      check per symbol per month (see `reconcile_with_statement`). It is
      NOT used for trade-level matching.

2. The order-history XLSX (e.g. "2026 - 02 - 15.xlsx") is a raw fill log
   (Entered/Filled time, Symbol w/ month code, Buy/Sell, Qty, Filled Price,
   Order Status, Commission, Contract Exp Date). This DOES have prices, so
   it is the source used for trade-level matching. Two important caveats
   found in the sample file, which will apply to any future export of the
   same kind:
     a. The export is a ROLLING WINDOW (in the sample: ~13 months ending at
        the export date). Trades that closed after the export's cutoff, or
        whose OPENING fill happened before the window started, will not
        reconcile correctly. This module surfaces the file's actual
        covered date range so gaps are visible rather than silently wrong.
     b. There is NO column identifying which of the 18 systems generated a
        given order (`Group Name` only holds bracket/OCO order-linkage IDs,
        not strategy names). When two systems trade the same underlying
        contract concurrently (e.g. several GC systems all trading MGC),
        a naive global FIFO reconstruction of "the" position will net
        unrelated systems' fills against each other and produce nonsense
        P&L. This module does NOT do that. Instead it matches fills
        directly against each system's theoretical trades (see
        `match_theoretical_to_actual`), leaving genuinely ambiguous or
        contested fills for human review rather than guessing.

3. Because there is no strategy tag on fills, and because sizing can be
   scaled at runtime in the dashboard (Tab 1), matching uses the actual
   `n_contracts` recorded on each theoretical trade row (not the system's
   ReadMe default), and always operates in raw TradeStation dollars.

─────────────────────────────────────────────────────────────────────────────
ASSUMPTIONS THAT ARE MY BEST JUDGMENT, NOT VERIFIED FACTS  (flagged per your
preference — check these if anything looks off)
─────────────────────────────────────────────────────────────────────────────
  * CONTRACT_SPECS tick sizes / point values below are standard CME/COMEX/
    NYMEX specs as commonly published; they are not re-verified against a
    live exchange spec sheet in this session. Confirm against your broker's
    contract specifications before trusting $-slippage numbers.
  * Default tolerances: date ±1 business day, price ±2 ticks (as you
    specified). These are configurable via `MatchConfig`.
  * "Manual override" classification (a matched pair with price/date outside
    tolerance but same symbol+qty+side+rough timing) is a heuristic guess at
    trades you exited early/late — since no override log exists yet, these
    are ALWAYS presented for your confirmation/edit, never silently assumed.
"""

from __future__ import annotations

import csv
import os
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import openpyxl


# ═════════════════════════════════════════════════════════════════════════
# CONTRACT SPECS  (see docstring caveat above — verify against your broker)
# ═════════════════════════════════════════════════════════════════════════

CONTRACT_SPECS = {
    "ES":  {"tick": 0.25, "tick_value": 12.50, "point_value": 50.0},
    "MES": {"tick": 0.25, "tick_value": 1.25,  "point_value": 5.0},
    "NQ":  {"tick": 0.25, "tick_value": 5.00,  "point_value": 20.0},
    "MNQ": {"tick": 0.25, "tick_value": 0.50,  "point_value": 2.0},
    "GC":  {"tick": 0.10, "tick_value": 10.00, "point_value": 100.0},
    "MGC": {"tick": 0.10, "tick_value": 1.00,  "point_value": 10.0},
    "CL":  {"tick": 0.01, "tick_value": 10.00, "point_value": 1000.0},
    "MCL": {"tick": 0.01, "tick_value": 1.00,  "point_value": 100.0},
}

# Per-root "this is ordinary slippage, not a deliberate override" price
# tolerance, in POINTS. Used only for the tight/auto-matched stage of
# match_theoretical_to_actual(), and only when the fill is still within the
# normal TIME tolerance (cfg.date_tol_minutes) — i.e. the logic Andrea laid out
# reviewing July 2026 NQ ORB custom results: if the fill happened at
# essentially the same time as the signal, a big price gap is slippage (the
# market moved between signal and fill), not a discretionary choice, and
# should count as a confident match (flagged "high_slippage") rather than
# sit in manual override review. A fill outside the time window is still
# treated as a deliberate override needing a logged reason, regardless of
# price.
#
# All values below are Andrea-confirmed: NQ/MNQ 13pts and ES/MES 4pts from
# the 2026-07 NQ ORB review; GC/MGC 3pts ("2-3 points, upper end") and
# CL/MCL 0.15pts added after that. CL/MCL only backs one strategy she rarely
# trades manually, so this tolerance mainly guards against real slippage,
# not overrides, on that root.
SLIPPAGE_TOL_POINTS = {
    "NQ": 13.0, "MNQ": 13.0,
    "ES": 4.0, "MES": 4.0,
    "GC": 3.0, "MGC": 3.0,   # Andrea: "gold can have 2-3 points" -> upper end
    "CL": 0.15, "MCL": 0.15,  # Andrea: "oil should have 0.15 max"
}


# Theoretical trade timestamps come from the backtest platform's own local
# clock (European time), while order-history fill times are Chicago
# exchange time. Confirmed on real June 2026 data: EVERY successful match
# for NQ ORB custom (7 tight + 16 loose, all 23 of them) landed on an
# actual fill time 6h55m-7h05m AFTER its theoretical time -- a consistent
# ~7h offset, not noise. Before this constant existed, the matcher was
# only finding these by accident, via a 24h-wide date window (the old
# to happen to catch a same-day 7h gap; that same wide net is what let an
# unrelated fill 2 days away get claimed by mistake (the 06-18 false-match
# bug fixed separately). Making the offset explicit lets the date window
# stay tight and correct instead of "wide enough to fluke into working".
#
# Andrea's own caveat (2026-07): twice a year, during the ~2-week window
# where US and EU daylight-saving transitions don't align, the true gap
# to Chicago is 6h or 8h instead of 7h. This constant is a simple default,
# not DST-aware -- if matching quality drops around a US/EU DST transition
# week, this is the first thing to check/adjust.
THEO_TIMEZONE_OFFSET_HOURS = 7.0


def _slippage_tol_ticks(root: str, cfg: "MatchConfig") -> float:
    """Max ticks a same-time fill may deviate on price and still be treated
    as ordinary slippage (auto-matched, no manual review) rather than an
    override candidate. Falls back to cfg.tick_tol for any root without an
    entry in SLIPPAGE_TOL_POINTS."""
    tick = CONTRACT_SPECS[root]["tick"]
    pts = SLIPPAGE_TOL_POINTS.get(root)
    if pts is None:
        return cfg.tick_tol
    return pts / tick

MONTH_CODES = {"F": 1, "G": 2, "H": 3, "J": 4, "K": 5, "M": 6,
               "N": 7, "Q": 8, "U": 9, "V": 10, "X": 11, "Z": 12}

_SYMBOL_RE = re.compile(r"^([A-Z]{1,4})([FGHJKMNQUVXZ])(\d{2})$")


def parse_symbol(sym) -> Optional[tuple]:
    """'MESH26' -> ('MES', 3, 2026'). Returns None if it doesn't look like a
    standard futures order-ticket symbol (root + month code + 2-digit year)."""
    if not sym:
        return None
    s = str(sym).strip().upper()
    m = _SYMBOL_RE.match(s)
    if not m:
        return None
    root, mo, yr = m.groups()
    return root, MONTH_CODES[mo], 2000 + int(yr)


def _clean_money(val):
    """
    Robustly parses TradeStation's various numeric text formats:
    "$2,852.00", "($393.75)" (parens = negative), "0.14%", "n/a", "Open",
    "", None, or an already-numeric value. Returns float or np.nan.
    Needed because the CSV/ODS exports format money as text with $ / parens,
    unlike the xlsx exports which give plain numeric cells.
    """
    if val is None:
        return np.nan
    if isinstance(val, (int, float)):
        return float(val) if not (isinstance(val, float) and np.isnan(val)) else np.nan
    s = str(val).strip()
    if s == "" or s.lower() in ("n/a", "open", "--", "nan"):
        return np.nan
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = s.replace("$", "").replace(",", "").replace("%", "").strip()
    if s == "":
        return np.nan
    try:
        v = float(s)
    except ValueError:
        return np.nan
    return -v if neg else v


# ═════════════════════════════════════════════════════════════════════════
# CONFIG
# ═════════════════════════════════════════════════════════════════════════

@dataclass
class MatchConfig:
    # Tight time tolerance for the auto-match/anchor-calibration stage, in
    # MINUTES (not days). Theoretical entry/exit timestamps are candle-CLOSE
    # times (e.g. a 5-min-candle ORB signal at 09:50 fired sometime in
    # [09:45, 09:50)), so the real execution lags the printed theoretical
    # time by up to one candle -- confirmed on real June 2026 data at
    # 2.5-3.7 minutes for a handful of ES rev eur trades once fill times
    # were converted to Chicago time. Was 1 full day before -- that huge
    # window was silently doing double duty (masking a then-unfixed 7h
    # timezone gap between theoretical and fill timestamps); once that gap
    # was fixed at the source (see THEO_TIMEZONE_OFFSET_HOURS), the 24h
    # window became actively harmful: many unrelated same-day fills fell
    # inside it, breaking the "exactly one candidate" anchor-uniqueness
    # check in calibrate_price_offsets and causing MORE unmatched trades,
    # not fewer. Andrea: "we can reduce also the tolerance to a default of
    # 5 minutes, right now it is 1 business day, definitely too much."
    date_tol_minutes: float = 10.0
    tick_tol: int = 2             # ±2 ticks slippage tolerance
    override_max_date_days: int = 5   # beyond this, don't even offer as an "override" candidate
    # Raised from 40 -> 100 after the July 2026 NQ ORB review: a genuine
    # deliberate override (better entry price, same qty/side/rough time) can
    # be a much bigger price gap than ordinary slippage — Andrea's 07/02
    # entry override was 67 ticks (16.75pts on MNQ) and was being missed
    # entirely at the old 40-tick ceiling, landing in "unmatched" instead of
    # "pending review". This bucket ALWAYS requires her manual confirmation
    # before counting as a real override, so a generous ceiling here is low
    # risk — it just means more candidates get surfaced for review, never
    # silently auto-applied.
    override_max_ticks: int = 100     # beyond this price gap, treat as unrelated, not an override


# ═════════════════════════════════════════════════════════════════════════
# STEP 1 — README PARSING (self-contained copy of app.py's logic, so this
# module has no import-time dependency on app.py / Streamlit)
# ═════════════════════════════════════════════════════════════════════════

def parse_readme(readme_path: str) -> dict:
    """Same parsing rules as app.py's parse_readme(). Kept independent on
    purpose so this module can be run/tested standalone without importing
    the Streamlit app (which would execute st.set_page_config etc.)."""
    alloc_re = re.compile(r"(\d+)\s*([A-Z]{2,3})")
    entries = {}
    sort_idx = 0
    try:
        with open(readme_path, "r", encoding="utf-8", errors="ignore") as fh:
            for raw in fh:
                line = raw.strip()
                if ":" not in line:
                    continue
                colon = line.index(":")
                name_part = line[:colon].strip()
                alloc_part = line[colon + 1:].strip()
                if not name_part or not alloc_part:
                    continue
                clean = alloc_part.split(",")[0]
                matches = alloc_re.findall(clean)
                if not matches:
                    continue
                by_type = defaultdict(int)
                for qty_s, ctype in matches:
                    by_type[ctype] += int(qty_s)
                primary_ctype = matches[0][1]
                first_qty = int(matches[0][0])
                norm = name_part.lower().replace("-", " ")
                norm = re.sub(r"\s+", " ", norm).strip()
                entries[norm] = {
                    "display_name": name_part,
                    "default_n": first_qty,
                    "ctype": primary_ctype,
                    "sort_order": sort_idx,
                }
                sort_idx += 1
    except FileNotFoundError:
        pass
    return entries


_CTYPE_TO_FAMILY = {"ES": "ES", "MES": "ES", "NQ": "NQ", "MNQ": "NQ",
                    "GC": "GC", "MGC": "GC", "CL": "CL", "MCL": "CL"}


def match_stem_to_readme(stem: str, readme_entries: dict) -> dict:
    """Same word-subset matching approach as app.py's get_alloc()."""
    # Fallback filename-derived symbol (matches app.py's convention of
    # underscore-separated names, e.g. "ES_breakout_..."). Only used if the
    # ReadMe lookup below fails to resolve a contract type.
    fallback_symbol = stem.split("_")[0].upper()
    norm_stem = stem.lower().replace("__", " ").replace("_", " ")
    stem_words = set(norm_stem.split())
    best, best_len = None, 0
    for norm_name, entry in readme_entries.items():
        kw = set(norm_name.split())
        if kw.issubset(stem_words) and len(kw) > best_len:
            best_len, best = len(kw), entry
    if best:
        # Derive the display symbol family (ES/NQ/GC/CL) from the resolved
        # contract type rather than the filename — robust regardless of
        # whether the file is named "ES_..." or "ES rev eur - short.xlsx".
        symbol = _CTYPE_TO_FAMILY.get(best["ctype"], fallback_symbol)
        return {"symbol": symbol, "display_name": best["display_name"],
                "ctype": best["ctype"], "default_n": best["default_n"],
                "matched": True}
    return {"symbol": fallback_symbol, "display_name": f"⚠️ {stem}",
            "ctype": fallback_symbol, "default_n": 0, "matched": False}


# ═════════════════════════════════════════════════════════════════════════
# STEP 2 — THEORETICAL TRADES  (extends app.py's "Trades List" parsing to
# also capture entry/exit price, needed for slippage matching)
# ═════════════════════════════════════════════════════════════════════════

def _parse_ts_date(val):
    """Same Italian-locale date-fix logic as app.py's parse_ts_date."""
    if val is None:
        return None
    if isinstance(val, (datetime, pd.Timestamp)):
        ts = pd.Timestamp(val)
        try:
            ts = ts.replace(month=ts.day, day=ts.month)
        except ValueError:
            pass
        return ts
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%m/%d/%Y %H:%M", "%m/%d/%Y",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return pd.Timestamp(datetime.strptime(s, fmt))
        except ValueError:
            continue
    try:
        return pd.Timestamp(s)
    except Exception:
        return None


def extract_theoretical_trades(ws) -> pd.DataFrame:
    """
    Parses the 'Trades List' worksheet, same layout as app.py's
    extract_trades_raw(), PLUS entry_price / exit_price from col[4] ('Price'),
    which app.py's dashboard doesn't need but the matcher does.

    Columns (verified across the sample file):
      Entry row: col[0]=trade#, col[1]=direction ('Sell Short'/'Buy'...),
                 col[2]=entry_dt, col[4]=entry_price,
                 col[6]=n_contracts, col[7]=trade_pnl
      Exit row:  col[0]=None, col[1]=exit_type, col[2]=exit_dt,
                 col[4]=exit_price,
                 col[6]=trade P&L ($, full position), col[7]=cum P&L ($)
    """
    records = []
    header_found = False
    current = {}

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] == "#":
                header_found = True
            continue

        if (row[0] is not None and isinstance(row[0], (int, float))
                and float(row[0]) == int(float(row[0]))):
            current = {
                "trade_id":    int(row[0]),
                "direction":   str(row[1]).strip() if row[1] else "",
                "entry_date":  _parse_ts_date(row[2]),
                "entry_price": float(row[4]) if row[4] is not None else None,
                "n_contracts": int(row[6]) if row[6] is not None else 1,
            }
        elif row[0] is None and row[1] is not None and current:
            try:
                pnl = float(row[6]) if row[6] is not None else float("nan")
                cum_pnl = float(row[7]) if row[7] is not None else float("nan")
            except (TypeError, ValueError):
                current = {}
                continue
            records.append({
                **current,
                "exit_type":  str(row[1]).strip(),
                "exit_date":  _parse_ts_date(row[2]),
                "exit_price": float(row[4]) if row[4] is not None else None,
                "pnl":        pnl,
                "cum_pnl":    cum_pnl,
            })
            current = {}

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)
    df["pnl"] = pd.to_numeric(df["pnl"], errors="coerce")
    # Normalise direction to Long/Short for matching against fill sides
    df["side_dir"] = np.where(
        df["direction"].str.contains("Short", case=False, na=False), "Short", "Long")
    return (df.dropna(subset=["exit_date", "entry_date", "pnl"])
              .sort_values("exit_date").reset_index(drop=True))


def extract_theoretical_trades_csv(csv_path: str) -> pd.DataFrame:
    """
    Parses the "TradeStation Trades List" section out of a flat CSV
    performance-report export (TradeStation's non-Excel report format —
    used when the xlsx report generator isn't available). One CSV file
    contains ALL report sections stacked (Performance Summary, Trade
    Analysis, Trades List, Annual, Monthly, Weekly, Daily, Settings); this
    function locates the "#,Type,Date/Time,..." header and reads rows until
    the next bare section header (a line with no comma).

    Same 14-column layout as the xlsx "Trades List" sheet, but values are
    formatted as text ("$2,852.00", "($393.75)", "n/a") instead of numeric
    cells — cleaned via `_clean_money`. Dates are plain "DD/MM/YYYY HH:MM"
    text (validated via weekday-of-week check: 0 Saturday/Sunday fills
    under DD/MM, and 1662/2835 rows have day > 12 so are unambiguous) — no
    swap needed here, unlike the datetime-object columns in xlsx/ods order
    history.

    A still-open trade (exit shown as "Open"/"n/a") is skipped: there's no
    realized exit to match against yet.
    """
    with open(csv_path, encoding="utf-8", errors="ignore") as fh:
        lines = fh.read().split("\n")

    start = None
    for i, l in enumerate(lines):
        if l.startswith("#,Type,Date/Time"):
            start = i + 1
            break
    if start is None:
        return pd.DataFrame()

    records = []
    current = {}
    for raw in lines[start:]:
        if raw.strip() == "":
            continue
        if "," not in raw:
            break  # next bare section header (e.g. "Annual")
        row = next(csv.reader([raw]))
        row = row + [""] * max(0, 14 - len(row))

        col0 = row[0].strip()
        is_entry = col0 != "" and col0.replace(".", "", 1).isdigit()

        if is_entry:
            n_ct = _clean_money(row[6])
            current = {
                "trade_id":    int(float(col0)),
                "direction":   row[1].strip(),
                "entry_date":  _parse_ts_date(row[2].strip()),
                "entry_price": _clean_money(row[4]),
                "n_contracts": int(n_ct) if not np.isnan(n_ct) else 1,
            }
        elif col0 == "" and row[1].strip() and current:
            exit_date_raw = row[2].strip()
            if exit_date_raw in ("Open", ""):
                current = {}   # still-open position — no exit to match yet
                continue
            pnl = _clean_money(row[6])
            if np.isnan(pnl):
                current = {}
                continue
            records.append({
                **current,
                "exit_type":  row[1].strip(),
                "exit_date":  _parse_ts_date(exit_date_raw),
                "exit_price": _clean_money(row[4]),
                "pnl":        pnl,
                "cum_pnl":    _clean_money(row[7]),
            })
            current = {}

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)
    df["pnl"] = pd.to_numeric(df["pnl"], errors="coerce")
    df["side_dir"] = np.where(
        df["direction"].str.contains("Short", case=False, na=False), "Short", "Long")
    return (df.dropna(subset=["exit_date", "entry_date", "pnl"])
              .sort_values("exit_date").reset_index(drop=True))


def extract_csv_total_net_profit(csv_path: str) -> Optional[float]:
    """
    Reads the "Total Net Profit" (All Trades column) out of a flat CSV
    report's "Performance Summary" section. This is the CSV equivalent of
    the xlsx report's "Performance Summary" sheet, cell B5, which app.py
    uses as a cross-check that the Trades List's last cumulative P&L
    matches the report's own stated total. Returns None if not found so
    callers can skip the check gracefully rather than fail.
    """
    try:
        with open(csv_path, encoding="utf-8", errors="ignore") as fh:
            for line in fh:
                if line.startswith("Total Net Profit,"):
                    parts = line.split(",")
                    if len(parts) >= 2:
                        v = _clean_money(parts[1])
                        return None if np.isnan(v) else float(v)
                    return None
    except Exception:
        return None
    return None


def load_all_theoretical_trades(data_dir: str, readme_path: str):
    """
    Loads every system's theoretical trades from data_dir and returns
    (combined_df, excluded) where excluded is a list of
    {"stem", "reason"} dicts. Supports both .xlsx (openpyxl "Trades List"
    sheet) and .csv (flat report export) per system. Self-contained: does
    not import or depend on app.py.

    If BOTH an .xlsx and a .csv exist for the same system name (stem), the
    .csv is preferred and the .xlsx is skipped — CSV is the newer export
    path (xlsx report generation has been crashing TradeStation for this
    user), so it's assumed to be the more current/complete one.

    A system whose stem doesn't match anything in ReadMe.txt, OR whose
    ReadMe allocation is explicitly 0 contracts, is EXCLUDED entirely (not
    just left unmatched) — per Andrea: "if these contracts are not present
    in the readme file, or they are as zero, then they should be counted
    as... zero". Concretely this means their trades never enter theo_month,
    so they contribute nothing to Theoretical P&L / Unmatched Impact /
    Portfolio Totals, rather than showing up as large phantom "unmatched"
    P&L for a system she isn't actually running at that size. Found via
    the 2026-06 review: "ES ORB custom - long e short" (not in ReadMe at
    all) and "NQ bout regolare - short" (also not in ReadMe -- possibly a
    rename/typo of "NQ bout continua - short", which IS listed) were
    silently inflating Unmatched Impact by thousands of dollars because
    their unresolved contract_type could never match anything, not because
    of a real matching failure. `excluded` is returned (not just dropped
    silently) so the caller can surface exactly what got excluded and why
    -- this is a coarse stem-name heuristic (match_stem_to_readme), so a
    genuine rename could get excluded by mistake and should stay visible,
    not disappear without a trace.
    """
    readme_entries = parse_readme(readme_path)
    data_path = Path(data_dir)

    xlsx_files = {f.stem: f for f in sorted(data_path.glob("*.xlsx"))}
    csv_files = {f.stem: f for f in sorted(data_path.glob("*.csv"))}
    stems = sorted(set(xlsx_files) | set(csv_files))

    all_rows = []
    excluded = []
    for stem in stems:
        trades = pd.DataFrame()
        if stem in csv_files:
            trades = extract_theoretical_trades_csv(str(csv_files[stem]))
        elif stem in xlsx_files:
            try:
                wb = openpyxl.load_workbook(str(xlsx_files[stem]), data_only=True)
                if "Trades List" in wb.sheetnames:
                    trades = extract_theoretical_trades(wb["Trades List"])
            except Exception:
                trades = pd.DataFrame()
        if trades.empty:
            continue
        meta = match_stem_to_readme(stem, readme_entries)
        if not meta["matched"]:
            excluded.append({"stem": stem, "reason": "not found in ReadMe.txt",
                              "n_trades": len(trades)})
            continue
        if meta["default_n"] == 0:
            excluded.append({"stem": stem, "reason": "ReadMe allocation is 0 contracts",
                              "n_trades": len(trades)})
            continue
        trades = trades.copy()
        trades["system_stem"] = stem
        trades["system_name"] = meta["display_name"]
        trades["symbol"] = meta["symbol"]
        trades["contract_type"] = meta["ctype"]
        all_rows.append(trades)
    if not all_rows:
        return pd.DataFrame(), excluded
    combined = pd.concat(all_rows, ignore_index=True)
    combined["theo_id"] = combined.index

    # Shift theoretical entry/exit timestamps from the backtest platform's
    # local (European) clock to Chicago exchange time -- see
    # THEO_TIMEZONE_OFFSET_HOURS docstring above for the evidence. Original
    # values are kept in *_date_local so the UI can show both if needed.
    combined["entry_date_local"] = combined["entry_date"]
    combined["exit_date_local"] = combined["exit_date"]
    tz_shift = pd.Timedelta(hours=THEO_TIMEZONE_OFFSET_HOURS)
    combined["entry_date"] = combined["entry_date_local"] + tz_shift
    combined["exit_date"] = combined["exit_date_local"] + tz_shift

    return combined, excluded


# ═════════════════════════════════════════════════════════════════════════
# STEP 3 — ACTUAL FILLS  (order-history export)
# ═════════════════════════════════════════════════════════════════════════

def _parse_fill_dt(val):
    """
    Parses 'Entered' / 'Filled/Canceled' timestamps from order-history
    exports (xlsx or ods). Different export formats need DIFFERENT
    month/day handling — this isn't guessed, it's validated per-format by
    checking the resulting weekday distribution (futures markets are shut
    Saturdays and most of Sunday daytime, so a wrong day/month swap shows
    up immediately as a cluster of "Saturday fills", which is impossible):

      - native datetime/Timestamp objects (some xlsx rows arrive this way)
        -> ALWAYS swap month/day. Validated on the xlsx sample: unswapped,
        several rows land in impossible future months; swapped, they land
        correctly right after the neighbouring Feb 2026 rows.

      - text "MM/DD/YY HH:MM:SS AM/PM" (the xlsx export's plain-text rows)
        -> NO swap. Validated: already gives a clean Mon-Fri calendar, and
        many values have day > 12 so are unambiguous as MM/DD.

      - text "YYYY-MM-DD HH:MM:SS" (the ods export's rows) -> ALWAYS swap.
        Validated on the ods sample: unswapped gives 207 Saturday + 288
        Sunday "fills" (impossible) and dates running into Aug-Dec 2026
        (the future, relative to the file's own export date); swapped
        gives 0 Saturdays and 1 Sunday, and no future dates.

    If a new export format shows up, re-run this same weekday check before
    trusting either branch — don't assume it matches one of the above.
    """
    if isinstance(val, (datetime, pd.Timestamp)):
        ts = pd.Timestamp(val)
        try:
            ts = ts.replace(month=ts.day, day=ts.month)
        except ValueError:
            pass  # day > 12, no swap possible/needed -> already unambiguous
        return ts
    s = str(val).strip()
    try:
        return pd.Timestamp(datetime.strptime(s, "%m/%d/%y %I:%M:%S %p"))
    except Exception:
        pass
    try:
        ts = pd.Timestamp(datetime.strptime(s, "%Y-%m-%d %H:%M:%S"))
        try:
            return ts.replace(month=ts.day, day=ts.month)
        except ValueError:
            return ts
    except Exception:
        pass
    try:
        return pd.Timestamp(s)
    except Exception:
        return pd.NaT


_ORDER_HISTORY_EXTS = (".xlsx", ".ods")


def resolve_latest_export(path: str, extensions: tuple = _ORDER_HISTORY_EXTS) -> str:
    """
    If `path` points at a DIRECTORY, returns the most-recently-modified
    file in it matching `extensions` (by filesystem mtime) — so you can
    point the order-history field at a folder (e.g. trade_data/order_history/)
    once, and every time you drop in a fresh export, the tool automatically
    picks it up without you having to retype a filename. If `path` already
    points at a specific file, it's returned unchanged.

    Deliberately uses file mtime rather than parsing dates out of the
    filename — filename conventions vary/aren't guaranteed, but "most
    recently added to the folder" is a reliable, format-agnostic signal.
    """
    p = Path(path)
    if p.is_dir():
        candidates = [f for ext in extensions for f in p.glob(f"*{ext}")]
        if not candidates:
            return path  # nothing found; let the normal "not found" error surface
        return str(max(candidates, key=lambda f: f.stat().st_mtime))
    return path


def resolve_statement_pdf(path: str, year: int, month: int) -> Optional[str]:
    """
    Resolves the monthly-statement PDF for a given (year, month).

    Unlike order-history (a single file that keeps getting replaced with
    "the latest"), a statements folder accumulates one PDF PER MONTH, so
    "most recently modified" is the wrong heuristic here — reconciliation
    needs the file for the SELECTED month specifically, not just whichever
    file was dropped in last.

    TradeStation's monthly statement filenames observed so far follow
    ACCOUNTID_YYYYMM.pdf (e.g. "210PYV33_202601.pdf" for Jan 2026), so if
    `path` is a directory this looks for a filename containing that
    YYYYMM stamp first. Falls back to the single most-recently-modified
    PDF in the folder if no filename match is found (better than nothing,
    but the reconciliation numbers may then be for the wrong month — the
    caller should treat that fallback as best-effort, not a first-class
    match). Returns None if nothing usable is found.
    """
    if not path:
        return None
    p = Path(path)
    if p.is_file():
        return str(p)
    if not p.is_dir():
        return None
    pdfs = list(p.glob("*.pdf"))
    if not pdfs:
        return None
    stamp = f"{year:04d}{month:02d}"
    stamped = [f for f in pdfs if stamp in f.stem]
    if stamped:
        return str(stamped[0])
    # No filename match for this month — fall back to newest file in the
    # folder rather than silently skipping reconciliation.
    return str(max(pdfs, key=lambda f: f.stat().st_mtime))


def load_order_history(path: str) -> pd.DataFrame:
    """
    Supports both .xlsx (openpyxl) and .ods (pandas + odfpy) order-history
    exports — same column layout observed in both sample files. Uses
    openpyxl for xlsx (consistent with the rest of the codebase / app.py)
    and pandas' odf engine for ods (requires `pip install odfpy`).

    `path` may be a specific file OR a directory — see resolve_latest_export().
    """
    path = resolve_latest_export(path)
    ext = Path(path).suffix.lower()
    if ext == ".ods":
        return pd.read_excel(path, engine="odf", sheet_name=0)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    return pd.DataFrame(rows[1:], columns=rows[0])


def extract_fills(df_raw: pd.DataFrame):
    """
    Returns (in_scope_fills, out_of_scope_fills).
    in_scope = Filled orders whose symbol root is one of the 8 contract
    types traded by the 18 systems (ES/MES/NQ/MNQ/GC/MGC/CL/MCL).
    out_of_scope = Filled orders on other roots (e.g. RTY, HG/MHG) — kept
    for visibility in the review queue, never matched.
    """
    df = df_raw.copy()
    df = df[df["Order Status"].astype(str).str.strip() == "Filled"].copy()
    parsed = df["Symbol"].apply(parse_symbol)
    df["root"] = parsed.apply(lambda p: p[0] if p else None)
    df["contract_key"] = parsed
    df["fill_time"] = df["Filled/Canceled"].apply(_parse_fill_dt)
    df["qty_filled"] = pd.to_numeric(df["Qty Filled"], errors="coerce").fillna(0).astype(int)
    df["remaining_qty"] = df["qty_filled"]
    df["price"] = df["Filled Price"].apply(_clean_money)
    df["side"] = df["Type"].astype(str).str.strip()          # Buy / Sell
    df["commission"] = df["Commission"].apply(_clean_money).fillna(0.0)
    df = df.dropna(subset=["price", "fill_time"])
    df = df[df["qty_filled"] > 0].sort_values("fill_time").reset_index(drop=True)
    df["fill_id"] = df.index

    in_scope = df[df["root"].isin(CONTRACT_SPECS.keys())].copy()
    out_of_scope = df[~df["root"].isin(CONTRACT_SPECS.keys())].copy()
    return in_scope.reset_index(drop=True), out_of_scope.reset_index(drop=True)


def order_history_coverage(fills: pd.DataFrame) -> dict:
    """Reports the actual covered date range so gaps are visible, not silent."""
    if fills.empty:
        return {"min": None, "max": None, "n": 0}
    return {"min": fills["fill_time"].min(), "max": fills["fill_time"].max(),
            "n": len(fills)}


# ═════════════════════════════════════════════════════════════════════════
# STEP 4 — MATCHING ENGINE
# Matches each THEORETICAL trade's entry leg and exit leg directly against
# candidate fills (root + side + qty + date/price tolerance), rather than
# pre-netting a global FIFO position per symbol. This avoids corrupting P&L
# when multiple systems trade the same contract concurrently (see module
# docstring point 2b).
# ═════════════════════════════════════════════════════════════════════════

@dataclass
class MatchOutcome:
    matched: pd.DataFrame            # confidently matched theo<->actual
    override_candidates: pd.DataFrame  # matched but outside tight tolerance
    ambiguous: list                  # list of dicts: theo trades competing for the same fills
    unmatched_theoretical: pd.DataFrame
    unmatched_actual: pd.DataFrame
    out_of_scope_actual: pd.DataFrame


def _leg_sides(side_dir: str):
    """Long trade: enter via Buy, exit via Sell. Short trade: enter via Sell, exit via Buy."""
    return ("Buy", "Sell") if side_dir == "Long" else ("Sell", "Buy")


def _candidate_fills(fills: pd.DataFrame, root: str, side: str, target_time,
                      target_price: float, qty_needed: int,
                      date_tol_minutes: float, tick_tol: int) -> pd.DataFrame:
    tick = CONTRACT_SPECS[root]["tick"]
    price_tol = tick_tol * tick
    date_lo = target_time - pd.Timedelta(minutes=date_tol_minutes)
    date_hi = target_time + pd.Timedelta(minutes=date_tol_minutes)
    cand = fills[
        (fills["root"] == root) &
        (fills["side"] == side) &
        (fills["remaining_qty"] > 0) &
        (fills["fill_time"] >= date_lo) & (fills["fill_time"] <= date_hi) &
        ((fills["price"] - target_price).abs() <= price_tol)
    ].copy()
    if cand.empty:
        return cand
    cand["date_diff_days"] = (cand["fill_time"] - target_time).abs().dt.total_seconds() / 86400
    cand["price_diff_ticks"] = (cand["price"] - target_price).abs() / tick
    cand["score"] = cand["price_diff_ticks"] + cand["date_diff_days"]
    return cand.sort_values("score")


def _loose_candidate_fills(fills: pd.DataFrame, root: str, side: str, target_time,
                            target_price: float, cfg: MatchConfig) -> pd.DataFrame:
    """Wider search used only to propose 'possible manual override' matches."""
    tick = CONTRACT_SPECS[root]["tick"]
    price_tol = cfg.override_max_ticks * tick
    date_lo = target_time - pd.Timedelta(days=cfg.override_max_date_days)
    date_hi = target_time + pd.Timedelta(days=cfg.override_max_date_days)
    cand = fills[
        (fills["root"] == root) & (fills["side"] == side) &
        (fills["remaining_qty"] > 0) &
        (fills["fill_time"] >= date_lo) & (fills["fill_time"] <= date_hi) &
        ((fills["price"] - target_price).abs() <= price_tol)
    ].copy()
    if cand.empty:
        return cand
    cand["date_diff_days"] = (cand["fill_time"] - target_time).abs().dt.total_seconds() / 86400
    cand["price_diff_ticks"] = (cand["price"] - target_price).abs() / tick
    cand["score"] = cand["price_diff_ticks"] + cand["date_diff_days"] * 4  # date matters more here
    return cand.sort_values("score")


def _claim_qty(fills: pd.DataFrame, candidates: pd.DataFrame, qty_needed: int):
    """Greedily consumes remaining_qty from the best-scoring candidates until
    qty_needed is met. Returns (claimed_rows_list, achieved_qty, weighted_avg_price, weighted_commission)."""
    claimed = []
    got = 0
    for fid in candidates["fill_id"]:
        if got >= qty_needed:
            break
        idx = fills.index[fills["fill_id"] == fid][0]
        avail = fills.at[idx, "remaining_qty"]
        take = min(avail, qty_needed - got)
        if take <= 0:
            continue
        price = fills.at[idx, "price"]
        comm_per_ct = fills.at[idx, "commission"] / fills.at[idx, "qty_filled"] \
            if fills.at[idx, "qty_filled"] else 0.0
        claimed.append({"fill_id": fid, "qty": take, "price": price,
                         "commission": comm_per_ct * take,
                         "fill_time": fills.at[idx, "fill_time"]})
        fills.at[idx, "remaining_qty"] -= take
        got += take
    if got == 0:
        return claimed, 0, None, 0.0
    avg_price = sum(c["price"] * c["qty"] for c in claimed) / got
    total_comm = sum(c["commission"] for c in claimed)
    return claimed, got, avg_price, total_comm


def calibrate_price_offsets(theo_df: pd.DataFrame, fills: pd.DataFrame,
                            cfg: MatchConfig = MatchConfig(), min_anchors: int = 3) -> dict:
    """
    Detects price offsets between theoretical trade prices and actual fill
    prices, PER (system, contract-month) -- not just per system.

    WHY THIS EXISTS: on the sample data, "ES rev eur - short" was backtested
    on TradeStation's continuous symbol "@MESU26" (see its Settings sheet).
    TradeStation continuous contracts are back-adjusted by default, which
    shifts EVERY historical price by a constant (per contract-roll period)
    relative to the real, un-adjusted contract that actually gets traded.
    On the Jan 2026 sample this offset was a very tight cluster
    (111.75-114.0 points, median 112.25) across 9 independent anchor trades
    — far too consistent to be real slippage, and far too large for a ±2
    tick tolerance to ever match without correcting for it first.

    WHY PER-CONTRACT, NOT JUST PER-SYSTEM: back-adjustment only shifts data
    BEHIND the current front month -- a contract trading AS the current
    front month has ~0 adjustment. So a system whose traded contract rolls
    mid-month (e.g. MNQ M26 -> U26) has TWO different true offsets within
    the same month, not one. Found on Andrea's 2026-06 "NQ ORB custom"
    review: pre-roll (MNQM26) trades showed a consistent ~280-305pt gap,
    post-roll (MNQU26, the then-current front month) trades matched almost
    exactly (~4-40pt, ordinary slippage) -- confirmed against the order-
    history's own contract symbols, which showed the M26->U26 roll landing
    on 2026-06-17. A single whole-month offset averages these two real,
    different populations into one untrustworthy number (spread far past
    the 3-tick sanity threshold), so nothing got corrected and every
    pre-roll trade failed to match. Calibrating separately per contract
    key (root+month+year, from each anchor fill's own parsed symbol) fixes
    this without needing to know in advance when a roll happened.

    Anchors are unambiguous fills that share root+side+exact qty within
    the date tolerance, found independent of price, grouped by which
    specific contract (contract_key) they were filled on. Returns
    dict[stem] -> list of segments, each:
        {"contract_key", "offset", "date_min", "date_max",
         "n_anchors", "n_raw_anchors", "spread", "applied"}
    date_min/date_max are the observed fill-date range for that contract,
    used by _candidate_offsets_for() to guess which segment a given
    theoretical trade should try first. If a segment's anchors don't
    cluster tightly enough to trust, "applied" is False and its "offset"
    is 0.0 (safer default than guessing).
    """
    offsets_by_system: dict = {}
    if theo_df.empty:
        return offsets_by_system

    for stem, grp in theo_df.groupby("system_stem"):
        root = grp["contract_type"].iloc[0]
        if root not in CONTRACT_SPECS:
            continue
        samples_by_contract = defaultdict(list)
        for _, t in grp.iterrows():
            entry_side, exit_side = _leg_sides(t["side_dir"])
            for side, dt, px in [(entry_side, t["entry_date"], t["entry_price"]),
                                  (exit_side, t["exit_date"], t["exit_price"])]:
                date_lo = dt - pd.Timedelta(minutes=cfg.date_tol_minutes)
                date_hi = dt + pd.Timedelta(minutes=cfg.date_tol_minutes)
                cand = fills[(fills["root"] == root) & (fills["side"] == side) &
                             (fills["qty_filled"] == t["n_contracts"]) &
                             (fills["fill_time"] >= date_lo) & (fills["fill_time"] <= date_hi)]
                if len(cand) == 1:
                    row = cand.iloc[0]
                    ck = row["contract_key"]
                    if ck is None:
                        continue
                    samples_by_contract[ck].append((px - float(row["price"]), row["fill_time"]))

        segments = []
        for ck, pairs in samples_by_contract.items():
            samples = [p[0] for p in pairs]
            times = [p[1] for p in pairs]
            if len(samples) < min_anchors:
                continue
            tick = CONTRACT_SPECS[root]["tick"]
            # Robust to false anchors: a same-root/side/qty/date-window fill
            # can coincidentally be the ONLY candidate (so it looks like a
            # clean single-match "anchor") while actually belonging to an
            # unrelated trade — this happened on the ods sample (9 tight
            # anchors around +112, 1 spurious anchor at +196). A plain
            # std-dev over all samples gets dragged around by that one
            # outlier, so: take a preliminary median, drop anything more
            # than 5 ticks from it, then compute the real median/spread
            # from the surviving "inlier" anchors only.
            prelim_med = float(np.median(samples))
            inliers = [s for s in samples if abs(s - prelim_med) <= 5 * tick]
            if len(inliers) >= min_anchors:
                med = float(np.median(inliers))
                spread = float(np.std(inliers))
            else:
                med, spread, inliers = prelim_med, float(np.std(samples)), samples
            # Only trust it as a real "back-adjustment constant" if the
            # anchors agree tightly (within ~3 ticks of each other) AND the
            # offset itself is bigger than the tick tolerance would forgive
            # anyway (otherwise it's not worth adjusting for).
            applied = spread <= 3 * tick and abs(med) > cfg.tick_tol * tick
            segments.append({
                "contract_key": ck, "offset": med if applied else 0.0,
                "raw_offset": med,  # kept even when not "applied" -- see
                                    # _candidate_offsets_for(include_untrusted=True)
                "date_min": min(times), "date_max": max(times),
                "n_anchors": len(inliers), "n_raw_anchors": len(samples),
                "spread": spread, "applied": applied,
            })
        if segments:
            offsets_by_system[stem] = segments
    return offsets_by_system


def _candidate_offsets_for(stem: str, entry_date, price_offsets: Optional[dict],
                            include_untrusted: bool = False) -> list:
    """
    Ordered list of price offsets to try for one theoretical trade. A
    theoretical trade doesn't carry its own contract-month tag (continuous-
    contract backtests don't export that), so we can't know upfront which
    calibrated segment applies -- instead, try the segment whose observed
    fill-date range covers (or is nearest to) this trade's entry date
    first, since that's the most likely match, then fall through to any
    other calibrated segments for this system, then 0.0 as the final
    fallback (covers the case where the real fill predates/postdates every
    anchor, or no segment was trustworthy).

    `include_untrusted`: when True, also offers segments whose anchors
    didn't cluster tightly enough to auto-trust (applied=False) using their
    raw computed median instead of 0.0 -- intended ONLY for the loose/
    override-review stage, which already requires a human to confirm the
    match, so a noisier-but-plausible offset is worth surfacing there even
    though it's not safe to silently auto-match on. Found necessary on
    Andrea's 2026-06 MNQ M26->U26 roll: the pre-roll segment's anchors
    clustered around ~292pts but with ~14pt spread (real signal, just
    noisier than the usual <1pt continuous-offset case) -- too loose to
    trust blindly, but far too useful to discard rather than offer for
    review.
    """
    segments = (price_offsets or {}).get(stem, [])
    usable = [s for s in segments if s["applied"]]
    if include_untrusted:
        # Untrusted segments (spread too wide to auto-trust) still carry
        # their raw computed median under "raw_offset" -- only worth
        # offering here (the loose/override-review stage), never for a
        # silent tight auto-match.
        usable += [s for s in segments if not s["applied"] and s.get("n_anchors", 0) >= 1]
    if not usable:
        return [0.0]

    def _dist(seg):
        if seg["date_min"] <= entry_date <= seg["date_max"]:
            return pd.Timedelta(0)
        return min(abs(entry_date - seg["date_min"]), abs(entry_date - seg["date_max"]))

    ordered = sorted(usable, key=_dist)
    offs = [(s["offset"] if s["applied"] else s.get("raw_offset", 0.0)) for s in ordered]
    if 0.0 not in offs:
        offs.append(0.0)
    return offs


def _release_claims(fills: pd.DataFrame, claims: list) -> None:
    for c in claims:
        idx = fills.index[fills["fill_id"] == c["fill_id"]][0]
        fills.at[idx, "remaining_qty"] += c["qty"]


def _try_tight_leg_match(fills: pd.DataFrame, root: str, entry_side: str, exit_side: str,
                          entry_date, exit_date, entry_target: float, exit_target: float,
                          qty: int, cfg: MatchConfig, slip_tol_ticks: float):
    """One tight-tolerance attempt for both legs at a given price offset
    (already applied to entry_target/exit_target). Releases all claims and
    returns success=False if either leg alone doesn't reach qty, OR if only
    one leg reaches qty (a tight match requires BOTH legs to confirm --
    see the fix note in match_theoretical_to_actual for why a lone
    successful leg must also be released, not just the failed one)."""
    entry_cand = _candidate_fills(fills, root, entry_side, entry_date,
                                   entry_target, qty, cfg.date_tol_minutes, slip_tol_ticks)
    exit_cand = _candidate_fills(fills, root, exit_side, exit_date,
                                  exit_target, qty, cfg.date_tol_minutes, slip_tol_ticks)
    entry_claim, entry_got, entry_avg_px, entry_comm = _claim_qty(fills, entry_cand, qty)
    if entry_got < qty:
        _release_claims(fills, entry_claim)
        entry_claim, entry_got = [], 0
    exit_claim, exit_got, exit_avg_px, exit_comm = _claim_qty(fills, exit_cand, qty)
    if exit_got < qty:
        _release_claims(fills, exit_claim)
        exit_claim, exit_got = [], 0
    if entry_got == qty and exit_got != qty:
        _release_claims(fills, entry_claim)
        entry_claim, entry_got = [], 0
    if exit_got == qty and entry_got != qty:
        _release_claims(fills, exit_claim)
        exit_claim, exit_got = [], 0

    # Same chronological guard as the loose stage: reject (and release) a
    # pairing where the exit fill precedes the entry fill -- see the note
    # in match_theoretical_to_actual's loose loop for the real-data case
    # that motivated this.
    if entry_got == qty and exit_got == qty:
        entry_fill_time = max(c["fill_time"] for c in entry_claim)
        exit_fill_time = min(c["fill_time"] for c in exit_claim)
        if exit_fill_time < entry_fill_time:
            _release_claims(fills, entry_claim + exit_claim)
            entry_claim, entry_got = [], 0
            exit_claim, exit_got = [], 0

    success = entry_got == qty and exit_got == qty
    return {
        "success": success,
        "entry_claim": entry_claim, "entry_avg_px": entry_avg_px, "entry_comm": entry_comm,
        "n_alt_entry": len(entry_cand) - len(entry_claim),
        "exit_claim": exit_claim, "exit_avg_px": exit_avg_px, "exit_comm": exit_comm,
        "n_alt_exit": len(exit_cand) - len(exit_claim),
    }


def match_theoretical_to_actual(theo_df: pd.DataFrame, fills: pd.DataFrame,
                                 cfg: MatchConfig = MatchConfig(),
                                 price_offsets: Optional[dict] = None) -> MatchOutcome:
    """
    Core matching loop. `fills` is copied internally (remaining_qty is
    decremented as claims happen on the internal copy only).
    """
    fills = fills.copy()
    matched_rows = []
    override_rows = []
    unmatched_theo_rows = []
    ambiguous = []

    if theo_df.empty:
        return MatchOutcome(pd.DataFrame(), pd.DataFrame(), [], theo_df,
                             fills, pd.DataFrame())

    # Process tightest, least-flexible trades first: sort by n_contracts desc
    # (larger positions have fewer plausible alternative fills) then by
    # entry date, to reduce arbitrary greedy bias.
    order = theo_df.sort_values(["n_contracts", "entry_date"], ascending=[False, True]).index

    for i in order:
        t = theo_df.loc[i]
        root = t["contract_type"]
        if root not in CONTRACT_SPECS:
            unmatched_theo_rows.append(i)
            continue
        entry_side, exit_side = _leg_sides(t["side_dir"])
        qty = int(t["n_contracts"])

        # Correct for a back-adjustment offset (continuous-contract backtest
        # prices vs. real traded-contract prices) if one was calibrated --
        # see calibrate_price_offsets(). A system can have MULTIPLE valid
        # offsets within the same month if its traded contract rolled, so
        # try each candidate in order (nearest-dated segment first) rather
        # than assuming a single system-wide constant.
        slip_tol_ticks = _slippage_tol_ticks(root, cfg)
        off_candidates = _candidate_offsets_for(t["system_stem"], t["entry_date"], price_offsets)

        tight_result, tight_off = None, 0.0
        for off in off_candidates:
            entry_target = t["entry_price"] - off
            exit_target = t["exit_price"] - off
            res = _try_tight_leg_match(fills, root, entry_side, exit_side,
                                        t["entry_date"], t["exit_date"],
                                        entry_target, exit_target, qty, cfg, slip_tol_ticks)
            if res["success"]:
                tight_result, tight_off = res, off
                break

        if tight_result is not None:
            off = tight_off
            entry_target = t["entry_price"] - off
            exit_target = t["exit_price"] - off
            entry_claim, entry_avg_px, entry_comm = (tight_result["entry_claim"],
                tight_result["entry_avg_px"], tight_result["entry_comm"])
            exit_claim, exit_avg_px, exit_comm = (tight_result["exit_claim"],
                tight_result["exit_avg_px"], tight_result["exit_comm"])
            pv = CONTRACT_SPECS[root]["point_value"]
            sign = 1 if t["side_dir"] == "Long" else -1
            actual_pnl = (exit_avg_px - entry_avg_px) * qty * pv * sign
            entry_slip_ticks = round((entry_avg_px - entry_target) / CONTRACT_SPECS[root]["tick"], 2)
            exit_slip_ticks = round((exit_avg_px - exit_target) / CONTRACT_SPECS[root]["tick"], 2)
            # "High slippage" = outside the ORIGINAL tight tolerance (cfg.tick_tol)
            # even though it's within this root's wider slippage band — i.e. a
            # same-time fill whose price moved more than the base tolerance,
            # worth a visual flag even though it didn't need manual review.
            high_slippage = (abs(entry_slip_ticks) > cfg.tick_tol or
                              abs(exit_slip_ticks) > cfg.tick_tol)
            matched_rows.append({
                "theo_id": t["theo_id"], "system_name": t["system_name"],
                "system_stem": t["system_stem"], "symbol": t["symbol"],
                "contract_type": root, "direction": t["side_dir"], "n_contracts": qty,
                "theo_entry_date": t["entry_date"], "theo_entry_price": t["entry_price"],
                "theo_exit_date": t["exit_date"], "theo_exit_price": t["exit_price"],
                "theo_pnl": t["pnl"],
                "actual_entry_time": min(c["fill_time"] for c in entry_claim),
                "actual_entry_price": entry_avg_px,
                "actual_exit_time": max(c["fill_time"] for c in exit_claim),
                "actual_exit_price": exit_avg_px,
                "actual_pnl_gross": actual_pnl,
                "commission": entry_comm + exit_comm,
                "actual_pnl_net": actual_pnl - (entry_comm + exit_comm),
                "entry_slippage_ticks": entry_slip_ticks,
                "exit_slippage_ticks": exit_slip_ticks,
                "high_slippage": high_slippage,
                "price_offset_applied": off,
                "had_alt_candidates": bool(tight_result["n_alt_entry"] > 0 or tight_result["n_alt_exit"] > 0),
            })
            continue

        loose_success, loose_off = False, 0.0
        le_claim = lx_claim = []
        le_px = lx_px = le_comm = lx_comm = 0.0
        # Loose stage always requires Andrea's manual confirmation, so it's
        # safe to also try noisier, not-fully-trusted offsets here (e.g. a
        # rolled contract whose anchors didn't cluster tightly enough for
        # the tight/auto-match stage) -- see _candidate_offsets_for().
        loose_off_candidates = _candidate_offsets_for(t["system_stem"], t["entry_date"],
                                                        price_offsets, include_untrusted=True)
        for off in loose_off_candidates:
            entry_target = t["entry_price"] - off
            exit_target = t["exit_price"] - off
            loose_entry = _loose_candidate_fills(fills, root, entry_side, t["entry_date"],
                                                  entry_target, cfg)
            loose_exit = _loose_candidate_fills(fills, root, exit_side, t["exit_date"],
                                                 exit_target, cfg)
            le_claim, le_got, le_px, le_comm = _claim_qty(fills, loose_entry, qty)
            lx_claim, lx_got, lx_px, lx_comm = _claim_qty(fills, loose_exit, qty)
            # Chronological sanity check: an exit fill can never occur before
            # its own entry fill. Without this guard, the loose stage's wide
            # date/price net can grab a fill that superficially matches on
            # root/side/price but belongs to an unrelated, earlier trade --
            # "stealing" it from whichever theoretical trade actually needed
            # it and leaving that other trade stranded in unmatched_actual.
            # Confirmed on real June 2026 data: NQ ORB custom's 06-18 09:30
            # trade was claiming an "exit" fill from 06-16 (two days before
            # its own entry fill of 06-22), which starved the true 06-16
            # trade of its legitimate exit and left it unmatched.
            if le_got == qty and lx_got == qty:
                entry_fill_time = max(c["fill_time"] for c in le_claim)
                exit_fill_time = min(c["fill_time"] for c in lx_claim)
                if exit_fill_time >= entry_fill_time:
                    loose_success, loose_off = True, off
                    break
            _release_claims(fills, le_claim + lx_claim)
            le_claim, lx_claim = [], []

        if loose_success:
            off = loose_off
            entry_target = t["entry_price"] - off
            exit_target = t["exit_price"] - off
            pv = CONTRACT_SPECS[root]["point_value"]
            sign = 1 if t["side_dir"] == "Long" else -1
            actual_pnl = (lx_px - le_px) * qty * pv * sign
            override_rows.append({
                "theo_id": t["theo_id"], "system_name": t["system_name"],
                "system_stem": t["system_stem"], "symbol": t["symbol"],
                "contract_type": root, "direction": t["side_dir"], "n_contracts": qty,
                "theo_entry_date": t["entry_date"], "theo_entry_price": t["entry_price"],
                "theo_exit_date": t["exit_date"], "theo_exit_price": t["exit_price"],
                "theo_pnl": t["pnl"],
                "actual_entry_time": min(c["fill_time"] for c in le_claim),
                "actual_entry_price": le_px,
                "actual_exit_time": max(c["fill_time"] for c in lx_claim),
                "actual_exit_price": lx_px,
                "actual_pnl_gross": actual_pnl,
                "commission": le_comm + lx_comm,
                "actual_pnl_net": actual_pnl - (le_comm + lx_comm),
                "entry_diff_ticks": round((le_px - entry_target) / CONTRACT_SPECS[root]["tick"], 2),
                "exit_diff_ticks": round((lx_px - exit_target) / CONTRACT_SPECS[root]["tick"], 2),
                "price_offset_applied": off,
                "override_reason": "",       # filled in by the user in the Streamlit UI
                "status": "pending_review",
            })
        else:
            unmatched_theo_rows.append(i)

    matched_df = pd.DataFrame(matched_rows)
    override_df = pd.DataFrame(override_rows)
    unmatched_theo_df = theo_df.loc[unmatched_theo_rows] if unmatched_theo_rows else pd.DataFrame()

    unmatched_actual = fills[fills["remaining_qty"] > 0].copy()

    return MatchOutcome(matched=matched_df, override_candidates=override_df,
                         ambiguous=ambiguous, unmatched_theoretical=unmatched_theo_df,
                         unmatched_actual=unmatched_actual, out_of_scope_actual=pd.DataFrame())


# STEP 5 — OVERRIDE LOG PERSISTENCE  (CSV, human-editable, survives restarts)
# ═════════════════════════════════════════════════════════════════════════

OVERRIDE_LOG_COLUMNS = [
    "theo_id", "system_name", "system_stem", "symbol", "n_contracts",
    "theo_entry_date", "theo_entry_price", "theo_exit_date", "theo_exit_price", "theo_pnl",
    "actual_entry_time", "actual_entry_price", "actual_exit_time", "actual_exit_price",
    "actual_pnl_net", "entry_diff_ticks", "exit_diff_ticks",
    "override_reason", "reviewed_by", "reviewed_at",
    # "override" = matched outside tight tolerance (slippage / early-late exit / re-entry).
    # "skipped" = signal was never taken at all (no execution exists to match) — logged
    # from the Unmatched Theoretical section so the reason survives restarts too.
    "match_type",
]


def load_override_log(path: str) -> pd.DataFrame:
    if os.path.exists(path):
        return pd.read_csv(path)
    return pd.DataFrame(columns=OVERRIDE_LOG_COLUMNS)


def save_override_entry(path: str, entry: dict) -> None:
    """Appends or updates (by theo_id) one override decision in the CSV log.
    `entry["match_type"]` should be "override" (matched but outside tight
    tolerance) or "skipped" (signal never taken — no execution to match)."""
    df = load_override_log(path)
    entry = {k: entry.get(k, "") for k in OVERRIDE_LOG_COLUMNS}
    for col in OVERRIDE_LOG_COLUMNS:  # tolerate older CSVs saved before a column existed
        if col not in df.columns:
            df[col] = ""
    if not df.empty and (df["theo_id"] == entry["theo_id"]).any():
        df.loc[df["theo_id"] == entry["theo_id"], list(entry.keys())] = list(entry.values())
    elif df.empty:
        df = pd.DataFrame([entry])
    else:
        df = pd.concat([df, pd.DataFrame([entry])], ignore_index=True)
    df.to_csv(path, index=False)


# ═════════════════════════════════════════════════════════════════════════
# STEP 6 — MONTHLY STATEMENT PDF  (reconciliation only, see docstring)
# ═════════════════════════════════════════════════════════════════════════

_PDF_LINE_RE = re.compile(
    r"^(\d{1,2}/\d{1,2}/\d{2})\s+F1\s+(\d+)\s+(\d+)\s+(.*?)\s+(\d{2})\s+"
    r"(P&L|FEE/COMM)\s+US\s+([\d,]+\.\d{2})$"
)

# x0 (left-edge, in PDF points) below which an amount is in the DEBIT
# column, at/above which it's in the CREDIT column. Measured directly off
# the June 2026 statement's header row ("DEBIT" header x0=426.4-447.4,
# "CREDIT" header x0=493.6-518.8) and cross-checked against every P&L/
# FEE line's actual amount position in that file: DEBIT amounts landed at
# x0 426.4-447.4, CREDIT amounts at x0 502.0-523.0 -- a wide, unambiguous
# gap, so 470 is a safe midpoint. Re-validate this against any statement
# with a visibly different page layout before trusting it blindly.
_DEBIT_CREDIT_X_SPLIT = 470.0


def _extract_statement_lines(pdf_path: str) -> list:
    """
    Extracts (top, text, amount_x0) for every text line on the statement,
    using word POSITIONS (not just concatenated text). This is required
    because DEBIT and CREDIT are separate table columns on the real PDF --
    a losing trade's amount is printed in the DEBIT column, a winning
    trade's in the CREDIT column, and BOTH render as a plain positive
    number with no minus sign or other marker in the text itself. Losing
    (DEBIT) amounts can only be told apart from winning (CREDIT) amounts by
    which column they're horizontally positioned in on the page --
    information plain `page.extract_text()` discards entirely. Using plain
    text here previously summed every trade (win AND loss) as positive,
    which inflated statement P&L by roughly 65x on the June 2026 statement
    (found: $104,217 naive sum vs. the statement's own stated "NET FUTURES
    PROFIT OR LOSS" of $1,591.25) -- do not revert to plain-text parsing.
    """
    import pdfplumber
    from collections import defaultdict
    lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            by_top = defaultdict(list)
            for w in words:
                by_top[round(w["top"], 1)].append(w)
            for top in sorted(by_top):
                ws = sorted(by_top[top], key=lambda w: w["x0"])
                text = " ".join(w["text"] for w in ws)
                amount_x0 = ws[-1]["x0"] if ws else None
                lines.append((top, text, amount_x0))
    return lines


def parse_monthly_statement_pdf(pdf_path: str) -> pd.DataFrame:
    """
    Extracts the daily P&L / FEE lines from a TradeStation monthly
    commodity statement. NOTE: this has NO per-trade prices — see module
    docstring. Used only for `reconcile_with_statement`.

    `amount` is SIGNED: negative if the line's amount fell in the DEBIT
    column (a loss / a fee), positive if in the CREDIT column (a gain /
    a rebate) -- see _extract_statement_lines() for why this requires
    word-position extraction rather than plain text.
    """
    recs = []
    for _top, ln, amount_x0 in _extract_statement_lines(pdf_path):
        m = _PDF_LINE_RE.match(ln.strip())
        if not m:
            continue
        date_s, qty_open, qty_close, desc, code, kind, amt_s = m.groups()
        amt = float(amt_s.replace(",", ""))
        if amount_x0 is not None and amount_x0 < _DEBIT_CREDIT_X_SPLIT:
            amt = -amt
        recs.append({
            "date": pd.Timestamp(datetime.strptime(date_s, "%m/%d/%y")),
            "qty_open": int(qty_open), "qty_close": int(qty_close),
            "description": desc.strip(), "kind": kind,
            "amount": amt,
        })
    return pd.DataFrame(recs)


# Maps the free-text "DESCRIPTION" field on the statement to a contract root.
# Built from the sample statements; extend if new product descriptions show up.
STATEMENT_DESC_TO_ROOT = {
    "CMX GOLD":        "GC",
    "CMX EMICR GOLD":  "MGC",
    "CME MCRO NSDQ":   "MNQ",
    "CME MICRO S&P":   "MES",
    "IMM EMINI NSDQ":  "NQ",
    "NYM LT CRUDE":    "CL",
}


def _desc_to_root(desc: str) -> Optional[str]:
    for key, root in STATEMENT_DESC_TO_ROOT.items():
        if key in desc.upper():
            return root
    return None


def reconcile_with_statement(pdf_df: pd.DataFrame, fills: pd.DataFrame,
                              year: int, month: int) -> pd.DataFrame:
    """
    Per-symbol total P&L and commission: statement total vs. reconstructed
    total from the order-history fills, for the given calendar month.
    This is a coarse sanity check (aggregated, no trade-level claim), useful
    precisely because the PDF can't support trade-level matching.
    """
    if pdf_df.empty:
        return pd.DataFrame()

    pdf_df = pdf_df.copy()
    pdf_df["root"] = pdf_df["description"].apply(_desc_to_root)
    stmt = pdf_df[(pdf_df["date"].dt.year == year) & (pdf_df["date"].dt.month == month)]
    stmt_pnl = stmt[stmt["kind"] == "P&L"].groupby("root")["amount"].sum()
    # Fee/commission lines are DEBITs on the statement, so `amount` here is
    # negative (see parse_monthly_statement_pdf). fee_by_root (from the
    # order-history "commission" column) is tracked as a positive cost by
    # convention on that side, so flip sign here to compare like-for-like.
    stmt_fee = -stmt[stmt["kind"] == "FEE/COMM"].groupby("root")["amount"].sum()

    f = fills[(fills["fill_time"].dt.year == year) & (fills["fill_time"].dt.month == month)]
    fee_by_root = f.groupby("root")["commission"].sum()

    rows = []
    all_roots = set(stmt_pnl.index) | set(stmt_fee.index)
    for root in sorted(all_roots):
        rows.append({
            "root": root,
            "statement_pnl": round(stmt_pnl.get(root, 0.0), 2),
            "statement_fee": round(stmt_fee.get(root, 0.0), 2),
            "fills_fee_reconstructed": round(fee_by_root.get(root, 0.0), 2),
            "fee_match": abs(stmt_fee.get(root, 0.0) - fee_by_root.get(root, 0.0)) < 1.0,
        })
    if rows:
        rows.append({
            "root": "TOTAL",
            "statement_pnl": round(sum(r["statement_pnl"] for r in rows), 2),
            "statement_fee": round(sum(r["statement_fee"] for r in rows), 2),
            "fills_fee_reconstructed": round(sum(r["fills_fee_reconstructed"] for r in rows), 2),
            "fee_match": all(r["fee_match"] for r in rows),
        })
    return pd.DataFrame(rows)


# ═════════════════════════════════════════════════════════════════════════
# STEP 7 — PER-SYSTEM DELTA BREAKDOWN
# ═════════════════════════════════════════════════════════════════════════

def compute_system_breakdown(matched_df: pd.DataFrame, override_df: pd.DataFrame,
                              unmatched_theo_df: pd.DataFrame,
                              sizing_ratios: Optional[dict] = None,
                              commission_rates: Optional[dict] = None) -> pd.DataFrame:
    """
    One row per system: theoretical P&L, actual P&L, and the delta split
    into commissions + slippage + overrides + unmatched.

    sizing_ratios / commission_rates (both keyed by system_stem) are
    OPTIONAL. When given, "Theoretical P&L" is recomputed to match exactly
    what the Portfolio tab (Tab 2) would show for these trades: raw pnl
    scaled by (current_n / default_n) from the live sizing panel, net of
    the sidebar's assumed commission -- i.e.
    sum((pnl_i - comm_per_contract * n_contracts_i) * scale) per system,
    same formula as app.py's build_net_equity(). commission_rates is a
    $/CONTRACT round-trip rate, not a flat per-trade cost (confirmed
    against Andrea's real statement: commissions scale with contract
    count) -- multiplied by each trade's own n_contracts, matching
    build_net_equity()'s fix for the same issue. Without sizing_ratios/
    commission_rates, "Theoretical P&L" is the raw, unscaled,
    gross-of-commission number straight from the trade files (not
    comparable to Tab 2's numbers).

    The Actual side is NEVER scaled by sizing_ratios — it's real executed
    dollars at whatever quantity was actually traded, not a hypothetical.

    COMMISSIONS ACCOUNTING (fixed 2026-07 after Andrea found the four delta
    components didn't sum to Delta -- a real $438 gap on June data): when
    commission_rates is given, theo_pnl is ALREADY net of an estimated
    commission (sidebar rate x n_contracts) applied across every trade —
    matched, override, AND unmatched. Meanwhile "Commissions ($)" used to
    separately sum each matched/override trade's REAL per-fill commission
    (a different number, at a different implied rate, covering only a
    SUBSET of trades) and get subtracted a second time in the reader's head
    when eyeballing whether the components add up. Per Andrea: theo P&L
    SHOULD be commission-adjusted (using a rate consistent with the real
    statement, not a guess), but the breakdown's "Commissions ($)" column
    must describe the SAME deduction already embedded in theo_pnl, not a
    second, inconsistent one -- otherwise Commissions + Slippage +
    Override Impact + Unmatched Impact will never equal Delta. Fixed by:
    Fixed decomposition (exact, verified algebraically so the four
    components always sum to Delta): slippage/override_impact/
    unmatched_impact are computed from GROSS P&L (pre-commission) on both
    sides, so they never touch commission at all. Commissions ($) then
    carries the ENTIRE commission story as one number: the difference
    between the REAL commission actually paid on matched+override trades
    (from each fill's own `commission` column) and the ESTIMATED
    commission assumed inside theo_pnl (sidebar rate x contracts, across
    matched+override+unmatched). This is "commission drag beyond what was
    assumed" -- e.g. negative if real fees ran higher than the sidebar
    estimate. Without commission_rates (toggle off), theo_pnl carries no
    commission assumption and Commissions ($) reduces to just the real
    commission paid (unchanged from before).
    """
    def _col_sum(df: pd.DataFrame, col: str) -> float:
        if df.empty or col not in df.columns:
            return 0.0
        return float(df[col].sum())

    def _theo_pnl_for(m: pd.DataFrame, o: pd.DataFrame, u: pd.DataFrame, stem: str):
        """Returns (theo_pnl_net, theo_pnl_gross, scale, comm_embedded).
        gross is the raw, unscaled, no-commission number, times scale.
        net additionally subtracts the sidebar-rate commission estimate
        (comm_embedded) if commission_rates is given. comm_embedded is
        returned separately since Commissions ($) below needs it on its
        own, not just folded into net."""
        raw = _col_sum(m, "theo_pnl") + _col_sum(o, "theo_pnl") + _col_sum(u, "pnl")
        scale = (sizing_ratios or {}).get(stem, 1.0)
        gross = raw * scale
        if not sizing_ratios and not commission_rates:
            return raw, raw, 1.0, 0.0
        comm  = (commission_rates or {}).get(stem, 0.0)
        total_contracts = _col_sum(m, "n_contracts") + _col_sum(o, "n_contracts") + _col_sum(u, "n_contracts")
        comm_embedded = comm * total_contracts * scale
        net = gross - comm_embedded
        return net, gross, scale, comm_embedded

    systems = set()
    for df in (matched_df, override_df, unmatched_theo_df):
        if not df.empty and "system_name" in df.columns:
            systems |= set(df["system_name"])

    rows = []
    for sysname in sorted(systems):
        m = matched_df[matched_df["system_name"] == sysname] if not matched_df.empty else pd.DataFrame()
        o = override_df[override_df["system_name"] == sysname] if not override_df.empty else pd.DataFrame()
        u = unmatched_theo_df[unmatched_theo_df["system_name"] == sysname] if not unmatched_theo_df.empty else pd.DataFrame()
        stem = (m["system_stem"].iloc[0] if not m.empty else
                (o["system_stem"].iloc[0] if not o.empty else
                 (u["system_stem"].iloc[0] if not u.empty else sysname)))

        theo_pnl, theo_pnl_gross, scale, comm_embedded = _theo_pnl_for(m, o, u, stem)
        actual_pnl = _col_sum(m, "actual_pnl_net") + _col_sum(o, "actual_pnl_net")

        # Gross-only components -- never touch commission, so they can't
        # double-count against the single Commissions ($) line below.
        # (theo side uses each trade's RAW theo_pnl/pnl x scale, not the
        # commission-netted theo_pnl, so slippage/override/unmatched are
        # pure price-gap effects.)
        slippage = _col_sum(m, "actual_pnl_gross") - _col_sum(m, "theo_pnl") * scale
        override_impact = _col_sum(o, "actual_pnl_gross") - _col_sum(o, "theo_pnl") * scale
        unmatched_impact = -_col_sum(u, "pnl") * scale  # theoretical P&L never realized

        # Commissions ($) = comm_embedded (estimated cost assumed inside
        # theo_pnl) MINUS real_comm (actual commission paid on matched+
        # override trades, from each fill's own commission column).
        # Verified algebraically: theo_pnl + Commissions($) + slippage +
        # override_impact + unmatched_impact == actual_pnl, exactly:
        #   slippage+override_impact+unmatched_impact = actual_gross - theo_gross
        #   theo_pnl = theo_gross - comm_embedded
        #   => theo_pnl + (S+O+U) = actual_gross - comm_embedded
        #   actual_pnl = actual_gross - real_comm  (by definition of _net)
        #   => need + (comm_embedded - real_comm) to reach actual_pnl. QED.
        # Sign convention: POSITIVE means you paid LESS real commission than
        # assumed (a tailwind vs. the estimate); NEGATIVE means real fees
        # ran higher than assumed (a drag).
        real_comm = _col_sum(m, "commission") + _col_sum(o, "commission")
        commissions = comm_embedded - real_comm

        rows.append({
            "System": sysname,
            "Theoretical P&L": round(theo_pnl, 2),
            "Actual P&L": round(actual_pnl, 2),
            "Delta": round(actual_pnl - theo_pnl, 2),
            "Commissions ($)": round(commissions, 2),
            # Renamed from "Slippage"/"Override Impact"/"Unmatched Impact" ->
            # explicit "($)" suffix after Andrea read GC donchian's -$1,466
            # and NQ bout continua's dollar figures as POINTS of slippage
            # (which would be enormous/impossible for those roots). These
            # are actual-vs-theoretical P&L dollar deltas (price gap x point
            # value x contracts), not a price/tick measure -- confirmed on
            # the real trades she pasted: GC donchian's entry/exit price
            # gaps were all 0.0-0.4pts (genuinely tiny, well inside
            # tolerance), but GC's $100/point value turns that into
            # hundreds of dollars per trade.
            "Slippage ($)": round(slippage, 2),
            "Override Impact ($)": round(override_impact, 2),
            "Unmatched Impact ($)": round(unmatched_impact, 2),
            "# Matched": len(m), "# Overrides (pending review)": len(o), "# Unmatched Theo": len(u),
        })
    return pd.DataFrame(rows)


def compute_portfolio_totals(breakdown_df: pd.DataFrame) -> dict:
    """
    Sums the per-system breakdown into a single portfolio-level row.
    IMPORTANT: this total is only meaningful once (most of) the 18 systems'
    trade files are actually loaded — with only 1-2 systems present it will
    understate the real portfolio gap, not reveal it.
    """
    if breakdown_df.empty:
        return {}
    numeric_cols = ["Theoretical P&L", "Actual P&L", "Delta", "Commissions ($)",
                     "Slippage ($)", "Override Impact ($)", "Unmatched Impact ($)",
                     "# Matched", "# Overrides (pending review)", "# Unmatched Theo"]
    return {c: round(float(breakdown_df[c].sum()), 2) for c in numeric_cols
            if c in breakdown_df.columns}


# ═════════════════════════════════════════════════════════════════════════
# TOP-LEVEL ORCHESTRATION
# ═════════════════════════════════════════════════════════════════════════

def run_matching_engine(strategy_dir: str, readme_path: str, order_history_path: str,
                         year: int, month: int, pdf_path: Optional[str] = None,
                         cfg: MatchConfig = MatchConfig(),
                         sizing_ratios: Optional[dict] = None,
                         commission_rates: Optional[dict] = None) -> dict:
    """
    Runs the full pipeline for one calendar month and returns a dict with
    everything the Streamlit tab needs to render.

    sizing_ratios / commission_rates: optional, keyed by system_stem — pass
    these (from app.py's live Tab 1 sizing panel + sidebar commission
    inputs) to make "Theoretical P&L" directly comparable to what the
    Portfolio tab shows for the same month. See compute_system_breakdown().
    """
    theo_all, excluded_systems = load_all_theoretical_trades(strategy_dir, readme_path)
    theo_month = theo_all[(theo_all["exit_date"].dt.year == year) &
                           (theo_all["exit_date"].dt.month == month)].copy() if not theo_all.empty else theo_all

    raw = load_order_history(order_history_path)
    in_scope, out_of_scope = extract_fills(raw)
    coverage = order_history_coverage(pd.concat([in_scope, out_of_scope]) if not out_of_scope.empty else in_scope)

    price_offsets = calibrate_price_offsets(theo_month, in_scope, cfg)
    outcome = match_theoretical_to_actual(theo_month, in_scope, cfg, price_offsets)
    breakdown = compute_system_breakdown(outcome.matched, outcome.override_candidates,
                                          outcome.unmatched_theoretical,
                                          sizing_ratios, commission_rates)
    portfolio_totals = compute_portfolio_totals(breakdown)

    recon = pd.DataFrame()
    resolved_pdf_path = resolve_statement_pdf(pdf_path, year, month) if pdf_path else None
    pdf_is_stamped_match = bool(
        resolved_pdf_path and f"{year:04d}{month:02d}" in Path(resolved_pdf_path).stem)
    if resolved_pdf_path and os.path.exists(resolved_pdf_path):
        pdf_df = parse_monthly_statement_pdf(resolved_pdf_path)
        recon = reconcile_with_statement(pdf_df, in_scope, year, month)

    out_scope_month = out_of_scope[(out_of_scope["fill_time"].dt.year == year) &
                                    (out_of_scope["fill_time"].dt.month == month)] if not out_of_scope.empty else out_of_scope

    # outcome.unmatched_actual is fills leftover from matching against the
    # FULL multi-month order-history pool (matching needs that wide pool so
    # a fill isn't missed just because it's a day or two outside the
    # selected month). But that means it's mostly fills from OTHER months
    # that were never going to match June's theoretical trades in the first
    # place -- showing all of them as "unmatched" for June is misleading
    # (on Andrea's June review this showed 4,241 fills, i.e. nearly the
    # entire 18-month file). Filter down to the selected month here, same
    # as out_scope_month above, so this section only shows fills that were
    # genuinely available to match in June and didn't.
    unmatched_actual_month = outcome.unmatched_actual[
        (outcome.unmatched_actual["fill_time"].dt.year == year) &
        (outcome.unmatched_actual["fill_time"].dt.month == month)
    ] if not outcome.unmatched_actual.empty else outcome.unmatched_actual

    return {
        "coverage": coverage,
        "price_offsets": price_offsets,
        "theo_month": theo_month,
        "matched": outcome.matched,
        "override_candidates": outcome.override_candidates,
        "unmatched_theoretical": outcome.unmatched_theoretical,
        "unmatched_actual": unmatched_actual_month,
        "resolved_pdf_path": resolved_pdf_path,
        "pdf_is_stamped_match": pdf_is_stamped_match,
        "out_of_scope_actual": out_scope_month,
        "breakdown": breakdown,
        "portfolio_totals": portfolio_totals,
        "n_systems_loaded": theo_all["system_stem"].nunique() if not theo_all.empty else 0,
        "excluded_systems": excluded_systems,
        "reconciliation": recon,
    }


if __name__ == "__main__":
    import sys
    strategy_dir = sys.argv[1] if len(sys.argv) > 1 else "uploads"
    readme_path = sys.argv[2] if len(sys.argv) > 2 else "uploads/ReadMe.txt"
    order_hist = sys.argv[3] if len(sys.argv) > 3 else "uploads/2026 - 02 - 15.xlsx"
    pdf = sys.argv[4] if len(sys.argv) > 4 else "uploads/210PYV33_202601.pdf"
    yr, mo = 2026, 1

    result = run_matching_engine(strategy_dir, readme_path, order_hist, yr, mo, pdf)
    print("Order history coverage:", result["coverage"])
    print("\nTheoretical trades this month:", len(result["theo_month"]))
    print("Matched:", len(result["matched"]))
    print("Override candidates (pending review):", len(result["override_candidates"]))
    print("Unmatched theoretical:", len(result["unmatched_theoretical"]))
    print("Unmatched actual fills:", len(result["unmatched_actual"]))
    print("\nBreakdown:\n", result["breakdown"])
